import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.ithalde_indirilecek_kdv import process_ithalde_indirilecek_kdv


class IthaldeIndirilecekKdvTest(unittest.TestCase):
    def test_grouping_distribution_and_exceptions(self):
        folder = Path(__file__).resolve().parents[1] / "outputs" / "ithalde_kdv_unit_test"
        folder.mkdir(parents=True, exist_ok=True)
        muavin_path = folder / "muavin.xlsx"
        report_path = folder / "ithalat.xlsx"
        self._write_muavin(muavin_path)
        self._write_import_report(report_path)

        status, output_path, message = process_ithalde_indirilecek_kdv(
            str(muavin_path),
            str(report_path),
            str(folder),
        )

        self.assertEqual("partial", status)
        self.assertIn("3 kayıt/beyanname", message)
        workbook = load_workbook(output_path, data_only=False)
        result = workbook["Sonuç"]
        exceptions = workbook["İnceleme Gerekenler"]
        control = workbook["Kontrol Özeti"]

        self.assertEqual(2, result.max_row - 2)
        self.assertEqual("INV1", result["D2"].value)
        self.assertEqual("5KG", result["H2"].value)
        self.assertEqual(50, result["J2"].value)
        self.assertEqual("INV2", result["D3"].value)
        self.assertEqual("5KG", result["H3"].value)
        self.assertEqual(50, result["J3"].value)
        self.assertEqual(1111111111, result["F2"].value)
        self.assertEqual("=J2/'Kontrol Özeti'!$B$2", result["I2"].value)
        self.assertEqual("dd.mm.yyyy", result["B2"].number_format)

        notes = [exceptions.cell(row, 11).value for row in range(2, exceptions.max_row + 1)]
        self.assertTrue(any("Miktar birimleri farklı" in note for note in notes))
        self.assertTrue(any("ithalat raporunda bulunamadı" in note for note in notes))
        self.assertTrue(any("beyanname numarası bulunamadı" in note for note in notes))
        exception_amounts = [
            exceptions.cell(row, 4).value
            for row in range(2, exceptions.max_row + 1)
            if exceptions.cell(row, 4).value is not None
        ]
        self.assertEqual(85, sum(exception_amounts))
        self.assertEqual("dd.mm.yyyy", exceptions["B2"].number_format)
        self.assertEqual("ELLE TAMAMLAMA GEREKİYOR", control["B10"].value)
        workbook.close()

    @staticmethod
    def _write_muavin(path):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet.append([
            "Belge tarihi",
            "Belge başlığı metni",
            "Şirket kodu para birimi değeri",
            "Metin",
        ])
        worksheet.append([date(2026, 6, 1), "VAKIFBANK İTHALAT KDV ÖD.", 100, "ÖDEME (DEC-A)"])
        worksheet.append([date(2026, 6, 2), "VAKIFBANK İTHALAT KDV ÖD.", 50, "ÖDEME (DEC-B)"])
        worksheet.append([date(2026, 6, 3), "VAKIFBANK İTHALAT KDV ÖD.", 25, "ÖDEME (DEC-C)"])
        worksheet.append([date(2026, 6, 4), "VAKIFBANK İTHALAT KDV ÖD.", 10, "BEYANNAME YOK"])
        worksheet.append([date(2026, 6, 30), "HAZİRAN 2026 KDV TAHAKKUK", -185, "KDV TAHAKKUKU"])
        workbook.save(path)

    @staticmethod
    def _write_import_report(path):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet.append([
            "Fatura numarası",
            "Tedarikçi adı",
            "Mal grubu tanımı",
            "Teslimat miktarı",
            "SAS ölçü birimi",
            "Gümrük beyanname numarası",
            "Gümrük beyanname tarihi",
        ])
        worksheet.append(["INV-1", "TEDARİKÇİ A", "HAM KAĞIT", 2, "KG", "DEC-A", 20260601])
        worksheet.append(["INV-1", "TEDARİKÇİ A", "HAM KAĞIT", 3, "KG", "DEC-A", 20260601])
        worksheet.append(["INV-2", "TEDARİKÇİ A", "HAM KAĞIT", 5, "KG", "DEC-A", 20260601])
        worksheet.append(["INV-3", "TEDARİKÇİ B", "YEDEK PARÇA", 1, "KG", "DEC-B", 20260602])
        worksheet.append(["INV-3", "TEDARİKÇİ B", "YEDEK PARÇA", 1, "ADT", "DEC-B", 20260602])
        worksheet.append(["RAPORDA-VAR", "TEDARİKÇİ D", "HAM KAĞIT", 20, "KG", "DEC-D", 20260605])
        workbook.save(path)


if __name__ == "__main__":
    unittest.main()
