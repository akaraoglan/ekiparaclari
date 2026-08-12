import io
import os
import unittest
import uuid
from contextlib import contextmanager
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

import app as app_module
from tools.ithaldeindirilecekfinal import process_ithaldeindirilecekfinal


@contextmanager
def _flat_workspace(root):
    root.mkdir(parents=True, exist_ok=True)
    existing_files = {path for path in root.iterdir() if path.is_file()}
    prefix = uuid.uuid4().hex
    try:
        yield str(root), prefix
    finally:
        for path in root.iterdir():
            if path.is_file() and path not in existing_files:
                path.unlink(missing_ok=True)


class IthaldeIndirilecekFinalTest(unittest.TestCase):
    TEST_ROOT = Path(
        os.environ.get(
            "EKIPARACLARI_TEST_TMPDIR",
            Path(__file__).resolve().parents[1] / "outputs",
        )
    )

    @classmethod
    def setUpClass(cls):
        cls.TEST_ROOT.mkdir(parents=True, exist_ok=True)

    def test_formulas_correction_and_workbook_preservation(self):
        with _flat_workspace(self.TEST_ROOT) as (folder, prefix):
            input_path = Path(folder) / f"{prefix}_input.xlsx"
            self._write_workbook(input_path)

            status, output_path, message = process_ithaldeindirilecekfinal(
                str(input_path),
                folder,
            )

            self.assertEqual("success", status)
            self.assertIn("3 satır işlendi", message)
            self.assertIn("1 satırın KDV matrahı düzeltildi", message)

            workbook = load_workbook(output_path, data_only=False)
            result = workbook["Sonuç"]
            other = workbook["Diğer Sayfa"]

            self.assertEqual(500, result["I2"].value)
            self.assertEqual(500, result["I3"].value)
            self.assertEqual(250, result["I4"].value)
            for row in range(2, 5):
                self.assertFalse(isinstance(result.cell(row, 9).value, str))
                self.assertFalse(isinstance(result.cell(row, 10).value, str))
                self.assertEqual(f"=I{row}*0.20", result.cell(row, 13).value)
                self.assertEqual(f"=ROUND(J{row}-M{row},2)", result.cell(row, 14).value)
                self.assertEqual("#,##0.00", result.cell(row, 13).number_format)
                self.assertEqual("#,##0.00", result.cell(row, 14).number_format)

            self.assertIsNone(result["M1"].value)
            self.assertIsNone(result["N1"].value)
            self.assertIsNone(result["M5"].value)
            self.assertIsNone(result["N5"].value)
            self.assertEqual("TOPLAM", result["I5"].value)
            self.assertEqual(250, result["J5"].value)
            self.assertEqual("Korunacak içerik", other["A1"].value)
            self.assertEqual("00FF0000", other["A1"].fill.fgColor.rgb)
            self.assertTrue(workbook.calculation.fullCalcOnLoad)
            self.assertTrue(workbook.calculation.forceFullCalc)
            self.assertEqual("auto", workbook.calculation.calcMode)
            workbook.close()

    def test_missing_result_sheet_returns_turkish_error(self):
        with _flat_workspace(self.TEST_ROOT) as (folder, prefix):
            input_path = Path(folder) / f"{prefix}_missing.xlsx"
            workbook = Workbook()
            workbook.active.title = "Data"
            workbook.save(input_path)
            workbook.close()

            status, output_path, message = process_ithaldeindirilecekfinal(
                str(input_path),
                folder,
            )

            self.assertEqual("error", status)
            self.assertIsNone(output_path)
            self.assertIn("'Sonuç' sayfası bulunamadı", message)

    def test_empty_result_sheet_returns_error(self):
        with _flat_workspace(self.TEST_ROOT) as (folder, prefix):
            input_path = Path(folder) / f"{prefix}_empty.xlsx"
            workbook = Workbook()
            workbook.active.title = "Sonuç"
            workbook.save(input_path)
            workbook.close()

            status, output_path, message = process_ithaldeindirilecekfinal(
                str(input_path),
                folder,
            )

            self.assertEqual("error", status)
            self.assertIsNone(output_path)
            self.assertIn("işlenecek veri bulunamadı", message)

    def test_invalid_workbook_returns_error(self):
        with _flat_workspace(self.TEST_ROOT) as (folder, prefix):
            input_path = Path(folder) / f"{prefix}_invalid.xlsx"
            input_path.write_text("Excel değil", encoding="utf-8")

            status, output_path, message = process_ithaldeindirilecekfinal(
                str(input_path),
                folder,
            )

            self.assertEqual("error", status)
            self.assertIsNone(output_path)
            self.assertIn("Excel dosyası açılamadı", message)

    def test_web_route_accepts_single_xlsx_and_downloads_result(self):
        with _flat_workspace(self.TEST_ROOT) as (folder, prefix):
            input_path = Path(folder) / f"{prefix}_web-input.xlsx"
            self._write_workbook(input_path)
            payload = input_path.read_bytes()

            app_module.app.config.update(TESTING=True)
            with patch.object(app_module, "UPLOAD_DIR", folder), patch.object(
                app_module,
                "OUTPUT_DIR",
                folder,
            ):
                response = app_module.app.test_client().post(
                    "/starwood/ithaldeindirilecekfinal",
                    data={"excel_file": (io.BytesIO(payload), "girdi.xlsx")},
                    content_type="multipart/form-data",
                )

            self.assertEqual(200, response.status_code)
            self.assertIn("attachment", response.headers["Content-Disposition"])
            self.assertIn("ithaldeindirilecekfinal_", response.headers["Content-Disposition"])
            downloaded = load_workbook(io.BytesIO(response.data), data_only=False)
            result = downloaded["Sonuç"]
            self.assertEqual("=I2*0.20", result["M2"].value)
            self.assertEqual("=ROUND(J2-M2,2)", result["N2"].value)
            self.assertEqual(500, result["I2"].value)
            self.assertEqual(250, result["J5"].value)
            downloaded.close()
            response.close()

    @staticmethod
    def _write_workbook(path):
        workbook = Workbook()
        result = workbook.active
        result.title = "Sonuç"
        result.append([
            "SIRA NO",
            "TARİH",
            "SERİ NO",
            "FATURA NO",
            "CARİ HESAP ADI",
            "VERGİ NO",
            "FATURA İZAHAT",
            "MİKTAR",
            "KDV MATRAHI",
            "KDV TUTARI",
            "GGB",
            "KDV DÖNEMİ",
        ])
        result.append([1, None, None, None, None, None, None, None, 500, 100])
        result.append([2, None, None, None, None, None, None, None, 499.95, 100])
        result.append([
            3,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "=J4/'Kontrol Özeti'!$B$2",
            50,
        ])
        result["I5"] = "TOPLAM"
        result["J5"] = "=SUM(J2:J4)"

        control = workbook.create_sheet("Kontrol Özeti")
        control["B2"] = 0.20
        other = workbook.create_sheet("Diğer Sayfa")
        other["A1"] = "Korunacak içerik"
        other["A1"].fill = PatternFill("solid", fgColor="FF0000")
        workbook.save(path)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
