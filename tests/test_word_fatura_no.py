import io
import os
import unittest
import uuid
from contextlib import contextmanager
import zipfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

import app as app_module
from tools.word_fatura_no import extract_invoice_numbers, word_invoices_to_excel


@contextmanager
def _flat_workspace(root: Path):
    root.mkdir(parents=True, exist_ok=True)
    existing_files = {path for path in root.iterdir() if path.is_file()}
    prefix = uuid.uuid4().hex
    try:
        yield str(root), prefix
    finally:
        for path in root.iterdir():
            if path.is_file() and path not in existing_files:
                path.unlink(missing_ok=True)


class WordFaturaNoTest(unittest.TestCase):
    TEST_ROOT = Path(
        os.environ.get(
            "EKIPARACLARI_TEST_TMPDIR",
            Path(__file__).resolve().parents[1] / "outputs",
        )
    )

    @classmethod
    def setUpClass(cls):
        cls.TEST_ROOT.mkdir(parents=True, exist_ok=True)

    def test_extracts_all_rows_from_target_table(self):
        with _flat_workspace(self.TEST_ROOT) as (folder, prefix):
            input_path = Path(folder) / f"{prefix}_STARWOOD - AKANSU.docx"
            self._write_document(input_path, 45)

            numbers = extract_invoice_numbers(str(input_path))

            self.assertEqual(45, len(numbers))
            self.assertEqual("SE22026000000001", numbers[0])
            self.assertEqual("SE22026000000045", numbers[-1])

    def test_creates_one_named_sheet_per_document(self):
        with _flat_workspace(self.TEST_ROOT) as (folder, prefix):
            first = Path(folder) / f"{prefix}_bir.docx"
            second = Path(folder) / f"{prefix}_iki.docx"
            self._write_document(first, 2)
            self._write_document(second, 1)

            status, output_path, message = word_invoices_to_excel(
                [(str(first), "STARWOOD - AKANSU.docx"), (str(second), "STARWOOD - AKANSU 2.docx")],
                folder,
            )

            self.assertEqual("success", status)
            self.assertIn("3 fatura numarası", message)
            workbook = load_workbook(output_path)
            self.assertEqual(["STARWOOD - AKANSU", "STARWOOD - AKANSU 2"], workbook.sheetnames)
            self.assertEqual("Fatura No", workbook["STARWOOD - AKANSU"]["A1"].value)
            self.assertEqual("SE22026000000002", workbook["STARWOOD - AKANSU"]["A3"].value)
            self.assertEqual("A2", workbook["STARWOOD - AKANSU"].freeze_panes)
            workbook.close()

    def test_web_route_downloads_a_single_workbook(self):
        with _flat_workspace(self.TEST_ROOT) as (folder, prefix):
            first = Path(folder) / f"{prefix}_bir.docx"
            second = Path(folder) / f"{prefix}_iki.docx"
            self._write_document(first, 2)
            self._write_document(second, 3)

            app_module.app.config.update(TESTING=True)
            with patch.object(app_module, "UPLOAD_DIR", folder), patch.object(
                app_module,
                "OUTPUT_DIR",
                folder,
            ):
                response = app_module.app.test_client().post(
                    "/starwood/word-fatura-nolari",
                    data={
                        "word_files": [
                            (io.BytesIO(first.read_bytes()), "bir.docx"),
                            (io.BytesIO(second.read_bytes()), "iki.docx"),
                        ]
                    },
                    content_type="multipart/form-data",
                )

            self.assertEqual(200, response.status_code)
            self.assertIn("attachment", response.headers["Content-Disposition"])
            workbook = load_workbook(io.BytesIO(response.data))
            self.assertEqual(["bir", "iki"], workbook.sheetnames)
            self.assertEqual(3, workbook["bir"].max_row)
            self.assertEqual(4, workbook["iki"].max_row)
            workbook.close()
            response.close()

    @staticmethod
    def _write_document(path: Path, invoice_count: int) -> None:
        rows = [
            WordFaturaNoTest._xml_row(
                ["FATURANIN TARİHİ", "FATURANIN NOSU", "KDV DAHİL TUTAR"]
            )
        ]
        for index in range(1, invoice_count + 1):
            rows.append(
                WordFaturaNoTest._xml_row(
                    ["01.01.2026", f"SE220260{index:08d}", "1.000,00"]
                )
            )

        document_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
            f"<w:body><w:tbl>{''.join(rows)}</w:tbl><w:sectPr/></w:body></w:document>"
        )
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("word/document.xml", document_xml.encode("utf-8"))

    @staticmethod
    def _xml_row(values: list[str]) -> str:
        cells = "".join(f"<w:tc><w:p><w:r><w:t>{value}</w:t></w:r></w:p></w:tc>" for value in values)
        return f"<w:tr>{cells}</w:tr>"


if __name__ == "__main__":
    unittest.main()
