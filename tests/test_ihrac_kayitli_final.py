import io
import os
import unittest
import uuid
from contextlib import contextmanager
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import app as app_module
from tools.ihrac_kayitli_final import process_ihrac_kayitli_final


@contextmanager
def _flat_workspace(root):
    root.mkdir(parents=True, exist_ok=True)
    existing_files = {path for path in root.iterdir() if path.is_file()}
    prefix = uuid.uuid4().hex
    try:
        yield str(root), prefix
    finally:
        for path in root.iterdir():
            if path.is_file() and path not in existing_files:
                path.unlink(missing_ok=True)


class IhracKayitliFinalTest(unittest.TestCase):
    TEST_ROOT = Path(
        os.environ.get(
            "EKIPARACLARI_TEST_TMPDIR",
            Path(__file__).resolve().parents[1] / "outputs",
        )
    )

    @classmethod
    def setUpClass(cls):
        cls.TEST_ROOT.mkdir(parents=True, exist_ok=True)

    def test_all_differences_and_total_are_reconciled(self):
        with _flat_workspace(self.TEST_ROOT) as (folder, prefix):
            input_path = Path(folder) / f"{prefix}_input.xlsx"
            self._write_workbook(input_path)

            status, output_path, message = process_ihrac_kayitli_final(str(input_path), folder)

            self.assertEqual("success", status)
            self.assertIn("3 satır işlendi", message)
            workbook = load_workbook(output_path, data_only=False)
            result = workbook["Sonuc"]
            expected_bases = [Decimal("500.00"), Decimal("250.00"), Decimal("1364411.80")]
            actual_total = Decimal("0")
            for row, expected in zip(range(5, 8), expected_bases):
                actual = Decimal(str(result.cell(row, 11).value)).quantize(Decimal("0.01"))
                kdv = Decimal(str(result.cell(row, 12).value)).quantize(Decimal("0.01"))
                difference = kdv - actual * Decimal("0.20")
                self.assertEqual(expected, actual)
                self.assertEqual(Decimal("0.00"), difference)
                actual_total += actual

            total_kdv = Decimal(str(result["M8"].value)).quantize(Decimal("0.01"))
            self.assertEqual(total_kdv / Decimal("0.20"), actual_total)
            self.assertEqual("TOPLAM", result["K8"].value)
            self.assertEqual(float(actual_total), result["L8"].value)
            self.assertEqual(15, result.max_column)
            self.assertEqual(4.7109375, result.column_dimensions["B"].width)
            self.assertEqual(64.7109375, result.column_dimensions["F"].width)
            self.assertEqual(140.25, result.row_dimensions[4].height)
            self.assertEqual("Arial", result["F5"].font.name)
            self.assertEqual(9, result["F5"].font.sz)
            self.assertEqual("left", result["F5"].alignment.horizontal)
            self.assertEqual("right", result["K5"].alignment.horizontal)
            self.assertEqual("landscape", result.page_setup.orientation)
            workbook.close()

    def test_inconsistent_kdv_total_returns_error(self):
        with _flat_workspace(self.TEST_ROOT) as (folder, prefix):
            input_path = Path(folder) / f"{prefix}_bad-total.xlsx"
            self._write_workbook(input_path, total_kdv=Decimal("273032.35"))

            status, output_path, message = process_ihrac_kayitli_final(str(input_path), folder)

            self.assertEqual("error", status)
            self.assertIsNone(output_path)
            self.assertIn("eşleşmiyor", message)

    def test_web_route_downloads_final_workbook(self):
        with _flat_workspace(self.TEST_ROOT) as (folder, prefix):
            input_path = Path(folder) / f"{prefix}_web.xlsx"
            self._write_workbook(input_path)
            payload = input_path.read_bytes()

            app_module.app.config.update(TESTING=True)
            with patch.object(app_module, "UPLOAD_DIR", folder), patch.object(
                app_module,
                "OUTPUT_DIR",
                folder,
            ):
                response = app_module.app.test_client().post(
                    "/starwood/ihrac-kayitli-final",
                    data={"excel_file": (io.BytesIO(payload), "girdi.xlsx")},
                    content_type="multipart/form-data",
                )

            self.assertEqual(200, response.status_code)
            self.assertIn("attachment", response.headers["Content-Disposition"])
            self.assertIn("ihrac_kayitli_final_", response.headers["Content-Disposition"])
            downloaded = load_workbook(io.BytesIO(response.data), data_only=False)
            self.assertEqual(15, downloaded["Sonuc"].max_column)
            self.assertEqual(1364411.8, downloaded["Sonuc"]["K7"].value)
            self.assertEqual(1365161.8, downloaded["Sonuc"]["L8"].value)
            self.assertEqual(273032.36, downloaded["Sonuc"]["M8"].value)
            downloaded.close()
            response.close()

    @staticmethod
    def _write_workbook(path, total_kdv=Decimal("273032.36")):
        workbook = Workbook()
        result = workbook.active
        result.title = "Sonuc"
        result["K4"] = "Malın Kdv Hariç Tutarı "
        result["L4"] = " Malın Kdv’si "
        result.append([])
        result["K5"] = 499.98
        result["L5"] = 100
        result["K6"] = 250.02
        result["L6"] = 50
        result["K7"] = 1364411.86
        result["L7"] = 272882.36
        result["K8"] = "TOPLAM"
        result["L8"] = float(total_kdv)
        workbook.save(path)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
