import io
import os
import unittest
import uuid
from contextlib import contextmanager
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import app as app_module
from tools.en_yuksek_alis_excel import (
    OUTPUT_SHEET_NAME,
    _format_tax_no,
    create_highest_purchase_workbook,
    read_top_forest_suppliers,
)


@contextmanager
def _flat_workspace(root: Path):
    root.mkdir(parents=True, exist_ok=True)
    existing_files = {path for path in root.iterdir() if path.is_file()}
    prefix = uuid.uuid4().hex
    try:
        yield str(root), prefix
    finally:
        for path in root.iterdir():
            if path.is_file() and path not in existing_files:
                path.unlink(missing_ok=True)


class EnYuksekAlisExcelTest(unittest.TestCase):
    TEST_ROOT = Path(
        os.environ.get(
            "EKIPARACLARI_TEST_TMPDIR",
            Path(__file__).resolve().parents[1] / "outputs",
        )
    )

    @classmethod
    def setUpClass(cls):
        cls.TEST_ROOT.mkdir(parents=True, exist_ok=True)

    def test_filters_groups_and_ranks_by_highest_single_invoice(self):
        with _flat_workspace(self.TEST_ROOT) as (folder, prefix):
            excel_path = Path(folder) / f"{prefix}.xlsx"
            self._write_excel(excel_path)

            suppliers = read_top_forest_suppliers(str(excel_path))

            self.assertEqual(["B ORMAN", "A ORMAN"], [item["name"] for item in suppliers])
            self.assertEqual(2, suppliers[0]["count"])
            self.assertEqual(156, suppliers[0]["amount"])
            self.assertEqual(240, suppliers[1]["amount"])

    def test_tax_number_is_written_without_spaces_and_keeps_leading_zero(self):
        self.assertEqual("0690516627", _format_tax_no(690516627))
        self.assertEqual("2222222222", _format_tax_no("222 222 2222"))

    def test_adds_copy_ready_sheet_and_preserves_source(self):
        with _flat_workspace(self.TEST_ROOT) as (folder, prefix):
            excel_path = Path(folder) / f"{prefix}.xlsx"
            output_path = Path(folder) / f"{prefix}_out.xlsx"
            self._write_excel(excel_path, supplier_count=10)

            result = create_highest_purchase_workbook(str(excel_path), str(output_path))
            workbook = load_workbook(output_path, data_only=True)
            try:
                self.assertIn("Data", workbook.sheetnames)
                self.assertIn(OUTPUT_SHEET_NAME, workbook.sheetnames)
                sheet = workbook[OUTPUT_SHEET_NAME]
                self.assertEqual("SOYADI/ADI VEYA UNVANI", sheet["A1"].value)
                self.assertEqual("VERGİ", sheet["B1"].value)
                self.assertEqual("DAİRESİ", sheet["B2"].value)
                self.assertEqual("VERGİ", sheet["C1"].value)
                self.assertEqual("NUMARASI", sheet["C2"].value)
                self.assertEqual("B ORMAN", sheet["A3"].value)
                self.assertIsNone(sheet["B3"].value)
                self.assertEqual("2222222222", sheet["C3"].value)
                self.assertEqual(2, sheet["D3"].value)
                self.assertEqual(156, sheet["E3"].value)
                self.assertEqual("n", sheet["E3"].data_type)
                self.assertEqual("#,##0.00", sheet["E3"].number_format)
                self.assertEqual({"A1:A2", "D1:D2", "E1:E2"}, {str(r) for r in sheet.merged_cells.ranges})
                self.assertEqual(10, result["supplier_count"])
            finally:
                workbook.close()

    def test_web_route_downloads_updated_excel(self):
        with _flat_workspace(self.TEST_ROOT) as (folder, prefix):
            excel_path = Path(folder) / f"{prefix}.xlsx"
            self._write_excel(excel_path, supplier_count=10)

            app_module.app.config.update(TESTING=True)
            with patch.object(app_module, "UPLOAD_DIR", folder), patch.object(
                app_module,
                "OUTPUT_DIR",
                folder,
            ):
                response = app_module.app.test_client().post(
                    "/starwood/en-yuksek-mallar",
                    data={
                        "excel_file": (io.BytesIO(excel_path.read_bytes()), "trivat.xlsx"),
                    },
                    content_type="multipart/form-data",
                )

            self.assertEqual(200, response.status_code)
            self.assertIn("trivat.xlsx", response.headers["Content-Disposition"])
            self.assertTrue(response.data.startswith(b"PK"))
            response.close()

    @staticmethod
    def _write_excel(path: Path, supplier_count=2):
        workbook = Workbook()
        data = workbook.active
        data.title = "Data"
        data.append(["Ad1", "Ad2", "Vergi no.", "Net tutar", "KDV"])
        data.append(["B ORMAN", "", "2222222222", 120, 24])
        data.append(["B ORMAN", "", "2222222222", 10, 2])
        data.append(["A ORMAN", "", "1111111111", 90, 18])
        data.append(["A ORMAN", "", "1111111111", 110, 22])
        data.append(["TEK ORMAN", "", "3333333333", 1000, 200])
        data.append(["SIFIRLI ORMAN", "", 690516627, 1, 0.2])
        for index in range(3, supplier_count + 1):
            data.append([f"C{index} ORMAN", "", f"{index:010d}", 80 - index, 10])
            data.append([f"C{index} ORMAN", "", f"{index:010d}", 5, 1])
        workbook.save(path)
        workbook.close()
