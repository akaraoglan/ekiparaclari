import io
import os
import unittest
import uuid
import zipfile
from contextlib import contextmanager
from datetime import date
from pathlib import Path
from unittest.mock import patch
from xml.dom import minidom

from openpyxl import Workbook

import app as app_module
from tools.word_yevmiye_doldur import (
    fill_word_journal_table,
    process_word_journal_pairs,
    read_excel_records,
)


@contextmanager
def _flat_workspace(root: Path):
    root.mkdir(parents=True, exist_ok=True)
    existing_files = {path for path in root.iterdir() if path.is_file()}
    prefix = uuid.uuid4().hex
    try:
        yield str(root), prefix
    finally:
        for path in root.iterdir():
            if path.is_file() and path not in existing_files:
                path.unlink(missing_ok=True)


class WordYevmiyeDoldurTest(unittest.TestCase):
    TEST_ROOT = Path(
        os.environ.get(
            "EKIPARACLARI_TEST_TMPDIR",
            Path(__file__).resolve().parents[1] / "outputs",
        )
    )

    @classmethod
    def setUpClass(cls):
        cls.TEST_ROOT.mkdir(parents=True, exist_ok=True)

    def test_fills_matching_values_and_leaves_attention_rows_blank(self):
        with _flat_workspace(self.TEST_ROOT) as (folder, prefix):
            word_path = Path(folder) / f"{prefix}.docx"
            excel_path = Path(folder) / f"{prefix}.xlsx"
            output_path = Path(folder) / f"{prefix}_out.docx"
            self._write_document(word_path, ["FATURA1", "FATURA2", "FATURA3"])
            self._write_excel(
                excel_path,
                [
                    ("FATURA1", date(2026, 1, 7), date(2026, 1, 7), 34022),
                    ("FATURA1", date(2026, 1, 7), date(2026, 1, 7), 34022),
                    ("FATURA2", date(2026, 1, 8), date(2026, 1, 9), 45000),
                ],
            )

            result = fill_word_journal_table(str(word_path), str(excel_path), str(output_path))
            rows = self._read_rows(output_path)

            self.assertEqual(["07.01.2026", "34022"], rows["FATURA1"])
            self.assertEqual(["", "45000"], rows["FATURA2"])
            self.assertEqual(["", ""], rows["FATURA3"])
            self.assertEqual(["FATURA2", "FATURA3"], result["attention_invoices"])

    def test_read_only_excel_is_consumed_in_one_sequential_pass(self):
        class SequentialWorksheet:
            def iter_rows(self, values_only=False):
                self.assert_values_only = values_only
                yield ("Referans", "Kayıt tarihi", "Belge tarihi", "Yevmiye No")
                yield ("FATURA1", date(2026, 1, 7), date(2026, 1, 7), 34022)
                yield ("FATURA1", date(2026, 1, 7), date(2026, 1, 7), 34022)

            def cell(self, *args, **kwargs):
                raise AssertionError("Read-only sayfada hücre bazlı erişim kullanılmamalı.")

        class SequentialWorkbook:
            sheetnames = ["Data"]
            epoch = None

            def __init__(self):
                self.worksheet = SequentialWorksheet()
                self.closed = False

            def __getitem__(self, key):
                return self.worksheet

            def close(self):
                self.closed = True

        workbook = SequentialWorkbook()
        with patch("tools.word_yevmiye_doldur.load_workbook", return_value=workbook):
            records = read_excel_records("test.xlsx")

        self.assertEqual("07.01.2026", records["fatura1"]["record_date"])
        self.assertEqual("34022", records["fatura1"]["journal_no"])
        self.assertTrue(workbook.worksheet.assert_values_only)
        self.assertTrue(workbook.closed)

    def test_zip_marks_document_that_needs_attention(self):
        with _flat_workspace(self.TEST_ROOT) as (folder, prefix):
            word_path = Path(folder) / f"{prefix}.docx"
            excel_path = Path(folder) / f"{prefix}.xlsx"
            self._write_document(word_path, ["FATURA1", "FATURA2"])
            self._write_excel(
                excel_path,
                [("FATURA1", date(2026, 1, 7), date(2026, 1, 7), 34022)],
            )

            status, zip_path, message = process_word_journal_pairs(
                [(str(word_path), str(excel_path), "STARWOOD - AKANSU.docx")],
                folder,
            )

            self.assertEqual("partial", status)
            self.assertIn("FATURA2", message)
            with zipfile.ZipFile(zip_path) as archive:
                self.assertEqual(["STARWOOD - AKANSU (ilgilen).docx"], archive.namelist())

    def test_web_route_matches_multiple_files_by_stem(self):
        with _flat_workspace(self.TEST_ROOT) as (folder, prefix):
            word_a = Path(folder) / f"{prefix}_a.docx"
            word_b = Path(folder) / f"{prefix}_b.docx"
            excel_a = Path(folder) / f"{prefix}_a.xlsx"
            excel_b = Path(folder) / f"{prefix}_b.xlsx"
            self._write_document(word_a, ["A1"])
            self._write_document(word_b, ["B1"])
            self._write_excel(excel_a, [("A1", date(2026, 1, 7), date(2026, 1, 7), 1)])
            self._write_excel(excel_b, [("B1", date(2026, 1, 8), date(2026, 1, 8), 2)])

            app_module.app.config.update(TESTING=True)
            with patch.object(app_module, "UPLOAD_DIR", folder), patch.object(
                app_module,
                "OUTPUT_DIR",
                folder,
            ):
                response = app_module.app.test_client().post(
                    "/starwood/word-yevmiye-doldur",
                    data={
                        "word_files": [
                            (io.BytesIO(word_a.read_bytes()), "A.docx"),
                            (io.BytesIO(word_b.read_bytes()), "B.docx"),
                        ],
                        "excel_files": [
                            (io.BytesIO(excel_b.read_bytes()), "B.xlsx"),
                            (io.BytesIO(excel_a.read_bytes()), "A.xlsx"),
                        ],
                    },
                    content_type="multipart/form-data",
                )

            self.assertEqual(200, response.status_code)
            self.assertIn("attachment", response.headers["Content-Disposition"])
            with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
                self.assertEqual(["A.docx", "B.docx"], sorted(archive.namelist()))
            response.close()

    @staticmethod
    def _write_excel(path: Path, rows) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data"
        sheet.append(["Referans", "Kayıt tarihi", "Belge tarihi", "Yevmiye No"])
        for row in rows:
            sheet.append(row)
        workbook.save(path)
        workbook.close()

    @staticmethod
    def _write_document(path: Path, invoice_numbers) -> None:
        header = WordYevmiyeDoldurTest._xml_row(
            ["FATURANIN TARİHİ", "FATURANIN NOSU", "KDV DAHİL TUTAR", "YEVMİYE KAYIT TARİH", "MAHSUP FİŞ NO"]
        )
        rows = [header]
        for invoice_number in invoice_numbers:
            rows.append(
                WordYevmiyeDoldurTest._xml_row(
                    ["01.01.2026", invoice_number, "1.000,00", "", ""]
                )
            )
        document_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
            f"<w:body><w:tbl>{''.join(rows)}</w:tbl><w:sectPr/></w:body></w:document>"
        )
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("word/document.xml", document_xml.encode("utf-8"))

    @staticmethod
    def _xml_row(values) -> str:
        cells = []
        for value in values:
            text = f"<w:r><w:t>{value}</w:t></w:r>" if value else ""
            cells.append(f"<w:tc><w:p>{text}</w:p></w:tc>")
        return f"<w:tr>{''.join(cells)}</w:tr>"

    @staticmethod
    def _read_rows(path: Path) -> dict[str, list[str]]:
        with zipfile.ZipFile(path) as archive:
            document = minidom.parseString(archive.read("word/document.xml"))
        rows = {}
        table_rows = document.getElementsByTagName("w:tr")[1:]
        for row in table_rows:
            cells = [
                "".join(
                    node.firstChild.data
                    for node in cell.getElementsByTagName("w:t")
                    if node.firstChild is not None
                )
                for cell in row.getElementsByTagName("w:tc")
            ]
            rows[cells[1]] = [cells[3], cells[4]]
        return rows
