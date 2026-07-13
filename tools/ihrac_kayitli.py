import os
import re
import uuid
from collections import defaultdict
from decimal import Decimal, InvalidOperation, ROUND_DOWN, ROUND_HALF_UP

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side


KDV_ORANI = Decimal("0.20")
M3_BOLEN = Decimal("1000000000")

DETAIL_COLUMNS = {
    "reference": ["Referans"],
    "tax_base": ["KDV Matrahı"],
    "currency": ["Belge para birimi"],
    "quantity": ["Faturalanan miktar"],
    "thickness": ["Kalınlık"],
    "length": ["Boy"],
    "width": ["En"],
    "gtip": ["Gtip", "GTİP", "GTIP"],
}

SUMMARY_COLUMNS = {
    "date": ["Faturalama tarihi"],
    "reference": ["Referans"],
    "buyer": ["Fatura alıcısı adı"],
    "tax_no": ["Vergi Numarası"],
}

RESULT_HEADERS = [
    "Sıra No",
    "İhraç Kayıtlı Satış Faturasının Tarihi ",
    "İhraç Kayıtlı Satış Faturasının Serisi ",
    " İhraç Kayıtlı Satış Faturasının Sıra No’su",
    "Alıcının Adı Soyadı/Ünvanı ",
    "Alıcının Vergi Kimlik Numarası / T.C. Kimlik Numarası ",
    " Malın Cinsi ",
    "Malın Miktarı ",
    "Miktar Kodu",
    "Malın Kdv Hariç Tutarı ",
    " Malın Kdv’si ",
    "İhracatçı Tarafından Yurt Dışına Düzenlenen Satış Faturasının Tarihi (GÇB/ETGB-BGB Üzerindeki Bilgiler)",
    "GÇB/ETGB-BGB Tescil No",
    "GÇB/ETGB-BGB",
]

RESULT_WIDTHS = {
    "A": 6.29,
    "B": 12.29,
    "C": 21.86,
    "D": 20.14,
    "E": 26.00,
    "F": 71.86,
    "G": 26.14,
    "H": 19.71,
    "I": 11.71,
    "J": 17.57,
    "K": 20.00,
    "L": 17.14,
    "M": 15.43,
    "N": 21.29,
    "O": 15.29,
    "P": 9.14,
}

RESULT_FORMATS = {
    "B": "0",
    "C": "mm-dd-yy",
    "D": "@",
    "E": "0",
    "F": "@",
    "G": "@",
    "H": "0",
    "I": "#,##0.00",
    "J": "@",
    "K": "#,##0.00",
    "L": "#,##0.00",
    "M": "mm-dd-yy",
    "N": "@",
    "O": "@",
}


def process_ihrac_kayitli(detay_path: str, ozet_path: str, output_dir: str) -> tuple:
    """
    Detay ve ozet Excel dosyalarindan ihrac kayitli satis faturasi listesi olusturur.
    Doner: (status, output_path, message)
    """
    try:
        summary_rows = _read_summary_rows(ozet_path)
        detail_groups, detail_refs = _read_detail_groups(detay_path)

        output_rows = []
        missing_refs = []
        used_refs = set()
        foreign_currency_rows = 0

        for summary in summary_rows:
            reference = summary["reference"]
            groups = detail_groups.get(reference)
            if not groups:
                missing_refs.append(reference)
                continue

            used_refs.add(reference)
            for gtip, group in sorted(groups.items(), key=lambda item: item[0], reverse=True):
                is_try = group["currencies"] and all(cur == "TRY" for cur in group["currencies"])
                if is_try:
                    miktar = _truncate_two(group["m3"])
                    matrah = _round_two(group["tax_base"])
                    kdv = _round_two(group["tax_base"] * KDV_ORANI)
                else:
                    miktar = None
                    matrah = None
                    kdv = None
                    foreign_currency_rows += 1

                output_rows.append({
                    "date": summary["date"],
                    "reference": reference,
                    "buyer": summary["buyer"],
                    "tax_no": summary["tax_no"],
                    "miktar": miktar,
                    "matrah": matrah,
                    "kdv": kdv,
                    "gtip": gtip,
                })

        if not output_rows:
            return "error", None, "Özet dosyasındaki referanslar için Detay dosyasında eşleşen kayıt bulunamadı."

        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"ihrac_kayitli_sonuc_{uuid.uuid4().hex[:8]}.xlsx")
        _write_result(output_rows, output_path)

        extra_refs = sorted(detail_refs - {row["reference"] for row in summary_rows})
        warnings = []
        if missing_refs:
            warnings.append(f"Özet'te olup Detay'da bulunmayan referans: {', '.join(missing_refs)}")
        if extra_refs:
            warnings.append(f"Detay'da olup Özet'te bulunmayan referans: {', '.join(extra_refs)}")
        message = f"Tamamlandı. {len(output_rows)} satır oluşturuldu."
        if foreign_currency_rows:
            message = f"{message} {foreign_currency_rows} dövizli satırda I, K ve L boş bırakıldı."
        if warnings:
            return "partial", output_path, f"{message} {' | '.join(warnings)}"
        return "success", output_path, message

    except Exception as exc:
        return "error", None, f"Hata oluştu: {exc}"


def _read_summary_rows(path: str) -> list:
    ws = _load_data_sheet(path)
    cols = _resolve_columns(ws, SUMMARY_COLUMNS, "Özet")
    rows = []

    for row_num in range(2, ws.max_row + 1):
        reference = _clean_text(ws.cell(row_num, cols["reference"]).value)
        if not reference:
            continue
        rows.append({
            "date": ws.cell(row_num, cols["date"]).value,
            "reference": reference,
            "buyer": _clean_text(ws.cell(row_num, cols["buyer"]).value),
            "tax_no": _clean_text(ws.cell(row_num, cols["tax_no"]).value),
        })

    if not rows:
        raise ValueError("Özet dosyasında işlenecek referans bulunamadı.")
    return rows


def _read_detail_groups(path: str) -> tuple:
    ws = _load_data_sheet(path)
    cols = _resolve_columns(ws, DETAIL_COLUMNS, "Detay")
    groups = defaultdict(dict)
    detail_refs = set()

    for row_num in range(2, ws.max_row + 1):
        reference = _clean_text(ws.cell(row_num, cols["reference"]).value)
        if not reference:
            continue

        detail_refs.add(reference)
        gtip = _clean_text(ws.cell(row_num, cols["gtip"]).value) or "GTIP_YOK"
        group = groups[reference].setdefault(gtip, {
            "tax_base": Decimal("0"),
            "m3": Decimal("0"),
            "currencies": set(),
        })

        tax_base = _decimal_value(ws.cell(row_num, cols["tax_base"]).value, row_num, "KDV Matrahı")
        quantity = _decimal_value(ws.cell(row_num, cols["quantity"]).value, row_num, "Faturalanan miktar")
        thickness = _decimal_value(ws.cell(row_num, cols["thickness"]).value, row_num, "Kalınlık")
        length = _decimal_value(ws.cell(row_num, cols["length"]).value, row_num, "Boy")
        width = _decimal_value(ws.cell(row_num, cols["width"]).value, row_num, "En")
        currency = _clean_text(ws.cell(row_num, cols["currency"]).value).upper()

        group["tax_base"] += tax_base
        group["m3"] += quantity * thickness * length * width / M3_BOLEN
        if currency:
            group["currencies"].add(currency)

    if not detail_refs:
        raise ValueError("Detay dosyasında işlenecek referans bulunamadı.")
    return groups, detail_refs


def _write_result(rows: list, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sonuc"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name="Arial", size=14, bold=True)
    header_font = Font(name="Arial", size=10, bold=True)
    data_font = Font(name="Arial", size=9)
    total_font = Font(name="Arial", size=9, bold=True)

    for col, width in RESULT_WIDTHS.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[4].height = 140.25

    ws["H2"] = "İHRAÇ KAYITLI SATIŞ FATURASI LİSTESİ"
    ws["H2"].font = title_font
    ws["H2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["H2"].number_format = "@"

    for offset, header in enumerate(RESULT_HEADERS, start=2):
        cell = ws.cell(4, offset, header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        cell.number_format = RESULT_FORMATS.get(cell.column_letter, "General")

    for index, item in enumerate(rows, start=1):
        row_num = index + 4
        values = [
            index,
            item["date"],
            "",
            item["reference"],
            item["buyer"],
            item["tax_no"],
            _clean_text(item["gtip"]),
            _to_float(item["miktar"]),
            "MTQ",
            _to_float(item["matrah"]),
            _to_float(item["kdv"]),
            None,
            None,
            "GÇB",
        ]

        for offset, value in enumerate(values, start=2):
            cell = ws.cell(row_num, offset, value)
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            cell.number_format = RESULT_FORMATS.get(cell.column_letter, "General")
            if cell.column_letter == "H" and value:
                cell.quotePrefix = True

    total_row = len(rows) + 5
    ws.cell(total_row, 11, "TOPLAM")
    ws.cell(total_row, 12, f"=SUM(K5:K{total_row - 1})")
    for cell in (ws.cell(total_row, 11), ws.cell(total_row, 12)):
        cell.font = total_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.number_format = "#,##0.00"

    wb.save(output_path)


def _load_data_sheet(path: str):
    workbook = load_workbook(path, data_only=True)
    if "Data" in workbook.sheetnames:
        return workbook["Data"]
    return workbook[workbook.sheetnames[0]]


def _resolve_columns(ws, required: dict, label: str) -> dict:
    normalized_headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value is not None:
            normalized_headers[_normalize_header(value)] = col

    resolved = {}
    missing = []
    for key, aliases in required.items():
        column = None
        for alias in aliases:
            column = normalized_headers.get(_normalize_header(alias))
            if column:
                break
        if column:
            resolved[key] = column
        else:
            missing.append(aliases[0])

    if missing:
        raise ValueError(f"{label} dosyasında eksik kolon: {', '.join(missing)}")
    return resolved


def _normalize_header(value) -> str:
    text = str(value).strip().lower()
    text = text.translate(str.maketrans({
        "ı": "i",
        "İ": "i",
        "ğ": "g",
        "Ğ": "g",
        "ü": "u",
        "Ü": "u",
        "ş": "s",
        "Ş": "s",
        "ö": "o",
        "Ö": "o",
        "ç": "c",
        "Ç": "c",
    }))
    return re.sub(r"[^a-z0-9]+", "", text)


def _decimal_value(value, row_num: int, field_name: str) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip().replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"{row_num}. satırda {field_name} sayısal değil: {value}") from exc


def _clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _round_two(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _truncate_two(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_DOWN)


def _to_float(value):
    if value is None:
        return None
    return float(value)
