from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
import os
import re
import threading
import unicodedata
import uuid

import fitz
from openpyxl import load_workbook

from tools.common import secure_tr_filename, zip_files


DECLARATION_RE = re.compile(r"\d{8}(?:IM|EX|AN)\d{6,10}", re.IGNORECASE)
EXPECTED_HEADERS = {
    "A": "teslimat",
    "D": "gumruk beyanname numarasi",
    "E": "intrastat grubu",
    "G": "para birimi",
    "H": "tsl.mkt.",
    "I": "satinalma blg",
    "J": "fatura",
}
_RAPID_OCR_ENGINE = None
_OCR_LOCK = threading.Lock()


@dataclass(frozen=True)
class ImportRow:
    delivery: str
    declaration: str
    amount: Decimal
    currency: str
    delivery_quantity: Decimal
    purchase_document: str
    invoice: str
    tl_amount: Decimal


def _plain_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value) -> str:
    text = _plain_text(value).translate(str.maketrans({"ı": "i", "İ": "I"}))
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii").casefold()
    return re.sub(r"\s+", " ", text).strip()


def _normalize_declaration(value) -> str:
    return re.sub(r"[^A-Z0-9]", "", _plain_text(value).upper())


def _decimal_value(value, row_number: int, column: str, label: str) -> Decimal:
    if value is None or _plain_text(value) == "":
        raise ValueError(f"Excel'in {column}{row_number} hücresindeki {label} boş.")
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(
            f"Excel'in {column}{row_number} hücresindeki {label} sayısal değil: {value}"
        ) from exc


def load_import_rows(excel_path: str) -> dict[str, list[ImportRow]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        data_sheets = [ws for ws in workbook.worksheets if ws.title.casefold() == "data"]
        if data_sheets:
            worksheet = data_sheets[0]
        elif len(workbook.worksheets) == 1:
            worksheet = workbook.worksheets[0]
        else:
            raise ValueError("Excel dosyasında 'Data' sayfası bulunamadı.")

        for column, expected in EXPECTED_HEADERS.items():
            actual = _normalize_header(worksheet[f"{column}1"].value)
            if actual != expected:
                raise ValueError(
                    f"Excel'in {column}1 hücresinde '{EXPECTED_HEADERS[column]}' başlığı bekleniyor."
                )

        grouped: dict[str, list[ImportRow]] = defaultdict(list)
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=2, min_col=1, max_col=11, values_only=True),
            start=2,
        ):
            declaration = _normalize_declaration(values[3])
            if not declaration:
                continue
            if not DECLARATION_RE.fullmatch(declaration):
                raise ValueError(
                    f"Excel'in D{row_number} hücresindeki beyanname numarası geçersiz: {values[3]}"
                )
            currency = _plain_text(values[6]).upper()
            if not currency:
                raise ValueError(f"Excel'in G{row_number} hücresindeki para birimi boş.")
            grouped[declaration].append(
                ImportRow(
                    delivery=_plain_text(values[0]),
                    declaration=declaration,
                    amount=_decimal_value(values[4], row_number, "E", "tutar"),
                    currency=currency,
                    delivery_quantity=_decimal_value(
                        values[7],
                        row_number,
                        "H",
                        "teslimat miktarı",
                    ),
                    purchase_document=_plain_text(values[8]),
                    invoice=_plain_text(values[9]),
                    tl_amount=_decimal_value(values[10], row_number, "K", "TL tutarı"),
                )
            )

        if not grouped:
            raise ValueError("Excel dosyasında işlenebilecek beyanname kaydı bulunamadı.")

        for rows in grouped.values():
            rows.sort(key=lambda row: _delivery_sort_key(row.delivery))
        return dict(grouped)
    finally:
        workbook.close()


def _delivery_sort_key(value: str):
    digits = re.sub(r"\D", "", value)
    if digits and digits == value:
        return (0, int(digits), value)
    return (1, value.casefold(), value)


def format_tr_amount(amount: Decimal) -> str:
    western = f"{amount:,.2f}"
    return western.replace(",", "_").replace(".", ",").replace("_", ".")


def format_tr_quantity(quantity: Decimal) -> str:
    decimal_text = format(quantity, "f")
    integer, dot, fraction = decimal_text.partition(".")
    grouped = f"{int(integer):,}".replace(",", ".")
    fraction = fraction.rstrip("0")
    return f"{grouped},{fraction}" if dot and fraction else grouped


def _alnum(text: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", text.upper())


def _ocr_char_matches(left: str, right: str) -> bool:
    groups = (set("0OQD"), set("1IL"), set("2Z"), set("5S"), set("6G"), set("8B"))
    return left == right or any(left in group and right in group for group in groups)


def _find_known_declaration(text: str, known: set[str]) -> str | None:
    compact = _alnum(text)
    exact = [declaration for declaration in known if declaration in compact]
    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1:
        raise ValueError("PDF sayfasında birden fazla beyanname numarası bulundu.")

    best: tuple[float, str] | None = None
    tied = False
    for declaration in known:
        length = len(declaration)
        if len(compact) < length:
            continue
        for start in range(len(compact) - length + 1):
            candidate = compact[start : start + length]
            matching = sum(
                _ocr_char_matches(expected, actual)
                for expected, actual in zip(declaration, candidate)
            )
            score = matching / length
            if best is None or score > best[0]:
                best = (score, declaration)
                tied = False
            elif best and score == best[0] and declaration != best[1]:
                tied = True

    if best and best[0] >= 0.89 and not tied:
        return best[1]
    return None


def _extract_declaration(page, known: set[str]) -> str | None:
    native_text = page.get_text("text")
    declaration = _find_known_declaration(native_text, known)
    if declaration:
        return declaration

    rapidocr_available = True
    try:
        rapidocr_text = _rapidocr_text(page)
    except ImportError:
        rapidocr_available = False
        rapidocr_text = ""
    except Exception:
        rapidocr_text = ""

    declaration = _find_known_declaration(rapidocr_text, known)
    if declaration:
        return declaration

    try:
        ocr_options = {"language": "eng", "dpi": 300, "full": True}
        tessdata = _find_tessdata()
        if tessdata:
            ocr_options["tessdata"] = tessdata
        text_page = page.get_textpage_ocr(**ocr_options)
    except RuntimeError as exc:
        if rapidocr_available:
            return None
        raise ValueError(
            "PDF taranmış görüntüden oluşuyor ve OCR bileşenleri kullanılamıyor. "
            "Sunucuda 'pip install -r requirements.txt' komutunu çalıştırın."
        ) from exc
    return _find_known_declaration(page.get_text(textpage=text_page), known)


def _rapidocr_text(page) -> str:
    import numpy as np
    from rapidocr import RapidOCR

    global _RAPID_OCR_ENGINE
    pixmap = page.get_pixmap(
        matrix=fitz.Matrix(2, 2),
        colorspace=fitz.csRGB,
        alpha=False,
    )
    image = np.frombuffer(pixmap.samples, dtype=np.uint8).reshape(
        pixmap.height,
        pixmap.width,
        pixmap.n,
    )
    with _OCR_LOCK:
        if _RAPID_OCR_ENGINE is None:
            _RAPID_OCR_ENGINE = RapidOCR()
        result = _RAPID_OCR_ENGINE(image)
    return "\n".join(result.txts or ())


def _find_tessdata() -> str | None:
    candidates = [
        os.environ.get("TESSDATA_PREFIX"),
        "/usr/share/tesseract-ocr/5/tessdata",
        "/usr/share/tesseract-ocr/4.00/tessdata",
        "/usr/share/tessdata",
    ]
    for candidate in candidates:
        if candidate and os.path.isfile(os.path.join(candidate, "eng.traineddata")):
            return candidate
    try:
        return fitz.get_tessdata()
    except RuntimeError:
        return None


def _currency_totals(rows: list[ImportRow]) -> dict[str, Decimal]:
    totals: dict[str, Decimal] = defaultdict(Decimal)
    for row in rows:
        totals[row.currency] += row.amount
    return dict(sorted(totals.items()))


def _invoice_totals(rows: list[ImportRow]):
    totals: dict[tuple[str, str], Decimal] = defaultdict(Decimal)
    invoices = set()
    for row in rows:
        invoice = row.invoice or "FATURA YOK"
        invoices.add(invoice)
        totals[(invoice, row.currency)] += row.amount
    return invoices, sorted(totals.items(), key=lambda item: (item[0][0], item[0][1]))


def _insert_text(page, point, text, *, size=8, bold=False, align=0, box=None):
    font = "hebo" if bold else "helv"
    if box is None:
        page.insert_text(point, text, fontsize=size, fontname=font, color=(0, 0, 0), overlay=True)
    else:
        page.insert_textbox(
            box,
            text,
            fontsize=size,
            fontname=font,
            color=(0, 0, 0),
            align=align,
            overlay=True,
        )


def _draw_page_content(doc, page, declaration: str, rows: list[ImportRow]):
    width = page.rect.width
    height = page.rect.height
    margin = 28.0
    bottom = height - 24.0
    first_page_start = min(max(height * 0.735, 585.0), height - 190.0)

    page.draw_rect(
        fitz.Rect(18, first_page_start - 16, width - 18, height - 12),
        color=None,
        fill=(1, 1, 1),
        overlay=True,
    )
    current_page = page
    y = first_page_start

    def new_continuation_page():
        nonlocal current_page, y, bottom
        current_page = doc.new_page(width=width, height=height)
        y = 38.0
        bottom = height - 24.0
        _insert_text(
            current_page,
            (margin, y),
            f"BEYANNAME: {declaration} - DEVAM",
            size=9,
            bold=True,
        )
        y += 18.0

    def ensure_space(required: float):
        if y + required > bottom:
            new_continuation_page()

    for currency, total in _currency_totals(rows).items():
        ensure_space(14)
        _insert_text(
            current_page,
            (margin, y),
            f"GENEL TOPLAM: {format_tr_amount(total)} {currency}",
            size=9,
            bold=True,
        )
        y += 14.0

    invoices, invoice_totals = _invoice_totals(rows)
    if len(invoices) > 1:
        for (invoice, currency), total in invoice_totals:
            ensure_space(12)
            _insert_text(
                current_page,
                (margin + 8, y),
                f"FATURA {invoice}: {format_tr_amount(total)} {currency}",
                size=8,
            )
            y += 12.0

    y += 8.0
    columns = [
        ("SATINALMA BLG", 0.00, 0.21, 0),
        ("TUTAR", 0.21, 0.38, 2),
        ("PB", 0.38, 0.45, 0),
        ("TESLIMAT MKT.", 0.45, 0.61, 2),
        ("TESLIMAT", 0.61, 0.79, 0),
        ("TL TUTAR", 0.79, 1.00, 2),
    ]
    table_left = margin
    table_width = width - 2 * margin
    row_height = 13.0

    def draw_cell(text, start, end, align, *, size=7.8, bold=False):
        font = "hebo" if bold else "helv"
        left = table_left + table_width * start + 3
        right = table_left + table_width * end - 3
        text_width = fitz.get_text_length(text, fontname=font, fontsize=size)
        if align == 2:
            x = max(left, right - text_width)
        elif align == 1:
            x = max(left, left + (right - left - text_width) / 2)
        else:
            x = left
        _insert_text(current_page, (x, y), text, size=size, bold=bold)

    def draw_header():
        nonlocal y
        ensure_space(row_height * 2)
        header_rect = fitz.Rect(table_left, y - 9, table_left + table_width, y + 4)
        current_page.draw_rect(header_rect, color=(0.45, 0.45, 0.45), fill=(0.92, 0.92, 0.92))
        for title, start, end, align in columns:
            draw_cell(title, start, end, align, size=7.3, bold=True)
        y += row_height

    draw_header()
    for row in rows:
        if y + row_height > bottom:
            new_continuation_page()
            draw_header()
        values = (
            row.purchase_document,
            format_tr_amount(row.amount),
            row.currency,
            format_tr_quantity(row.delivery_quantity),
            row.delivery,
            f"{format_tr_amount(row.tl_amount)} TL",
        )
        for value, (_, start, end, align) in zip(values, columns):
            draw_cell(value, start, end, align)
        current_page.draw_line(
            fitz.Point(table_left, y + 4),
            fitz.Point(table_left + table_width, y + 4),
            color=(0.8, 0.8, 0.8),
            width=0.4,
        )
        y += row_height


def annotate_pdf(input_path: str, output_path: str, rows_by_declaration) -> tuple[int, list[str]]:
    source = fitz.open(input_path)
    output = fitz.open()
    matched_pages = 0
    warnings = []
    known = set(rows_by_declaration)
    try:
        for page_number in range(source.page_count):
            source_page = source[page_number]
            try:
                declaration = _extract_declaration(source_page, known)
            except ValueError as exc:
                warnings.append(f"Sayfa {page_number + 1}: {exc}")
                declaration = None

            output.insert_pdf(source, from_page=page_number, to_page=page_number)
            output_page = output[-1]
            if declaration:
                _draw_page_content(output, output_page, declaration, rows_by_declaration[declaration])
                matched_pages += 1
            else:
                warnings.append(
                    f"Sayfa {page_number + 1}: Excel ile eşleşen beyanname numarası bulunamadı."
                )

        if matched_pages == 0:
            ocr_error = next(
                (warning for warning in warnings if "Tesseract OCR" in warning),
                None,
            )
            if ocr_error:
                raise ValueError(ocr_error.split(": ", 1)[-1])
            raise ValueError("PDF'de Excel ile eşleşen beyanname numarası bulunamadı.")
        output.save(output_path, garbage=4, deflate=True)
    finally:
        output.close()
        source.close()
    return matched_pages, warnings


def process_declaration_pdfs(pdf_files, excel_path: str, output_dir: str):
    rows_by_declaration = load_import_rows(excel_path)
    os.makedirs(output_dir, exist_ok=True)
    output_paths = []
    all_warnings = []
    matched_pdf_count = 0

    for index, (pdf_path, original_name) in enumerate(pdf_files, start=1):
        stem = os.path.splitext(secure_tr_filename(original_name))[0]
        output_name = f"{stem}_doldurulmus_{uuid.uuid4().hex[:8]}.pdf"
        output_path = os.path.join(output_dir, output_name)
        try:
            matched_pages, warnings = annotate_pdf(pdf_path, output_path, rows_by_declaration)
        except Exception as exc:
            all_warnings.append(f"{original_name}: {exc}")
            continue
        matched_pdf_count += 1
        output_paths.append(output_path)
        all_warnings.extend(f"{original_name} - {warning}" for warning in warnings)

    if not output_paths:
        detail = all_warnings[0] if all_warnings else "Eşleşme bulunamadı."
        raise ValueError(detail)

    status = "partial" if all_warnings or matched_pdf_count != len(pdf_files) else "success"
    message = f"{matched_pdf_count}/{len(pdf_files)} PDF dosyası eşleştirilip dolduruldu."
    if all_warnings:
        message += f" {len(all_warnings)} uyarı oluştu."

    if len(output_paths) == 1:
        return status, output_paths[0], os.path.basename(output_paths[0]), message

    zip_name = f"beyanname_pdfleri_{uuid.uuid4().hex[:8]}.zip"
    zip_path = os.path.join(output_dir, zip_name)
    zip_files(output_paths, zip_path)
    return status, zip_path, zip_name, message
