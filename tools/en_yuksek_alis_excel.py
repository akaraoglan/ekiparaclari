import os
import re
import uuid
from collections import defaultdict
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from tools.common import secure_tr_filename
from tools.word_yevmiye_doldur import _normalize_text


EXCEL_HEADERS = {
    "name1": {"ad1"},
    "name2": {"ad2"},
    "tax_no": {"vergino", "verginumarasi"},
    "net": {"nettutar"},
    "vat": {"kdv"},
}
OUTPUT_SHEET_NAME = "En Yüksek 10"
OUTPUT_HEADERS = [
    "SOYADI/ADI VEYA UNVANI",
    "VERGİ DAİRESİ",
    "VERGİ NUMARASI",
    "SAYISI",
    "TUTAR",
]
ATTENTION_SUFFIX = re.compile(r"\s*\((?:ilgilen|en yüksek 10)\)\s*$", re.IGNORECASE)


def _resolve_columns(row_values) -> dict[str, int] | None:
    normalized = {
        _normalize_text(value): index
        for index, value in enumerate(row_values)
        if value is not None
    }
    resolved = {}
    for key, aliases in EXCEL_HEADERS.items():
        for alias in aliases:
            column = normalized.get(_normalize_text(alias))
            if column is not None:
                resolved[key] = column
                break
    return resolved if len(resolved) == len(EXCEL_HEADERS) else None


def _decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip().replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Sayısal değer okunamadı: {value}") from exc


def _format_tax_no(value) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) == 10:
        return f"{digits[:3]} {digits[3:6]} {digits[6:]}"
    return str(value or "").strip()


def _read_top_forest_suppliers_from_rows(rows, limit: int = 10) -> list[dict]:
    columns = None
    grouped = defaultdict(
        lambda: {
            "name": "",
            "tax_no": "",
            "count": 0,
            "net_total": Decimal("0"),
            "vat_total": Decimal("0"),
            "max_net": Decimal("0"),
        }
    )

    for row_number, row_values in enumerate(rows, start=1):
        if columns is None:
            if row_number <= 20:
                columns = _resolve_columns(row_values)
                if columns is not None:
                    continue
            if row_number >= 20:
                raise ValueError(
                    "Data sayfasında Ad1, Ad2, Vergi no., Net tutar ve KDV başlıkları bulunamadı."
                )
            continue

        name = " ".join(
            part
            for part in (
                str(row_values[columns["name1"]] or "").strip(),
                str(row_values[columns["name2"]] or "").strip(),
            )
            if part
        )
        if "orman" not in name.casefold():
            continue

        net = _decimal(row_values[columns["net"]])
        vat = _decimal(row_values[columns["vat"]])
        key = _normalize_text(name)
        item = grouped[key]
        item["count"] += 1
        item["net_total"] += net
        item["vat_total"] += vat
        if not item["name"] or net > item["max_net"]:
            item["name"] = name
            item["tax_no"] = _format_tax_no(row_values[columns["tax_no"]])
        item["max_net"] = max(item["max_net"], net)

    if columns is None:
        raise ValueError(
            "Data sayfasında Ad1, Ad2, Vergi no., Net tutar ve KDV başlıkları bulunamadı."
        )

    qualified = [item for item in grouped.values() if item["count"] >= 2]
    qualified.sort(key=lambda item: (-item["max_net"], _normalize_text(item["name"])))
    for item in qualified:
        item["amount"] = item["net_total"] + item["vat_total"]
    return qualified[:limit]


def read_top_forest_suppliers(excel_path: str, limit: int = 10) -> list[dict]:
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    try:
        if "Data" not in workbook.sheetnames:
            raise ValueError("Excel dosyasında Data sayfası bulunamadı.")
        return _read_top_forest_suppliers_from_rows(
            workbook["Data"].iter_rows(values_only=True),
            limit,
        )
    finally:
        workbook.close()


def _create_output_sheet(workbook, suppliers: list[dict]):
    if OUTPUT_SHEET_NAME in workbook.sheetnames:
        workbook.remove(workbook[OUTPUT_SHEET_NAME])
    worksheet = workbook.create_sheet(OUTPUT_SHEET_NAME)
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Calibri", size=11, color="1F1F1F")
    alternate_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_gray = Side(style="thin", color="B4C6E7")
    border = Border(bottom=thin_gray)

    for column, header in enumerate(OUTPUT_HEADERS, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index in range(2, 12):
        supplier_index = row_index - 2
        supplier = suppliers[supplier_index] if supplier_index < len(suppliers) else None
        values = [
            supplier["name"] if supplier else "",
            "",
            supplier["tax_no"] if supplier else "",
            supplier["count"] if supplier else None,
            float(supplier["amount"]) if supplier else None,
        ]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column, value=value)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left" if column in (1, 2) else "center",
                vertical="center",
            )
            if row_index % 2 == 1:
                cell.fill = alternate_fill

    worksheet.column_dimensions["A"].width = 48
    worksheet.column_dimensions["B"].width = 22
    worksheet.column_dimensions["C"].width = 20
    worksheet.column_dimensions["D"].width = 12
    worksheet.column_dimensions["E"].width = 20
    worksheet.row_dimensions[1].height = 32
    for row_index in range(2, 12):
        worksheet.row_dimensions[row_index].height = 22

    for row_index in range(2, 12):
        worksheet.cell(row=row_index, column=3).number_format = "@"
        worksheet.cell(row=row_index, column=4).number_format = "#,##0"
        worksheet.cell(row=row_index, column=5).number_format = "#,##0.00"
    worksheet.auto_filter.ref = "A1:E11"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.print_area = "A1:E11"
    worksheet.sheet_view.zoomScale = 90
    return worksheet


def create_highest_purchase_workbook(excel_path: str, output_path: str) -> dict:
    suppliers = read_top_forest_suppliers(excel_path)
    workbook = load_workbook(excel_path)
    try:
        _create_output_sheet(workbook, suppliers)
        workbook.save(output_path)
    finally:
        workbook.close()
    return {
        "supplier_count": len(suppliers),
        "needs_attention": len(suppliers) < 10,
        "suppliers": suppliers,
    }


def process_highest_purchase_workbook(
    excel_path: str,
    original_excel_name: str,
    output_dir: str,
) -> tuple[str, str, str, str]:
    os.makedirs(output_dir, exist_ok=True)
    stem = os.path.splitext(os.path.basename(original_excel_name))[0]
    clean_stem = ATTENTION_SUFFIX.sub("", stem).strip() or "sonuc"
    safe_name = secure_tr_filename(f"{clean_stem}.xlsx")
    physical_path = os.path.join(output_dir, f"{uuid.uuid4().hex[:8]}_{safe_name}")

    result = create_highest_purchase_workbook(excel_path, physical_path)
    if result["needs_attention"]:
        download_name = f"{clean_stem} (ilgilen).xlsx"
        message = (
            f"Yalnızca {result['supplier_count']} uygun cari bulundu; kalan satırlar boş bırakıldı."
        )
        return "partial", physical_path, message, download_name

    return (
        "success",
        physical_path,
        "En yüksek 10 mal alış carisi Excel'deki yeni sayfaya eklendi.",
        f"{clean_stem}.xlsx",
    )
