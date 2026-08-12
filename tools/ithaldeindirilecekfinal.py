import os
import uuid
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from openpyxl import load_workbook


KDV_RATE = Decimal("0.20")
MONEY_STEP = Decimal("0.01")
RESULT_SHEET = "Sonuç"
CONTROL_SHEET = "Kontrol Özeti"


def process_ithaldeindirilecekfinal(input_path: str, output_dir: str):
    """Sonuç sayfasına KDV hesaplama ve fark kontrolü sütunlarını ekler."""
    workbook = None
    values_workbook = None
    try:
        workbook = load_workbook(input_path, data_only=False)
        values_workbook = load_workbook(input_path, data_only=True)
    except Exception:
        if workbook is not None:
            workbook.close()
        if values_workbook is not None:
            values_workbook.close()
        return "error", None, "Excel dosyası açılamadı. Lütfen geçerli bir .xlsx dosyası yükleyin."

    try:
        if RESULT_SHEET not in workbook.sheetnames:
            return "error", None, "Excel dosyasında 'Sonuç' sayfası bulunamadı."

        worksheet = workbook[RESULT_SHEET]
        values_worksheet = values_workbook[RESULT_SHEET]
        last_data_row = _find_last_data_row(worksheet)
        total_row = _find_total_row(worksheet)
        if last_data_row < 2:
            return "error", None, "'Sonuç' sayfasında işlenecek veri bulunamadı."

        processed_count = 0
        corrected_count = 0
        total_base = Decimal("0")
        total_kdv = Decimal("0")

        for row in range(2, last_data_row + 1):
            kdv_cell = worksheet.cell(row, 10)
            kdv_amount = _as_decimal(kdv_cell.value)
            if kdv_amount is None:
                kdv_amount = _as_decimal(values_worksheet.cell(row, 10).value)
            if isinstance(kdv_cell.value, str) and kdv_cell.value.startswith("=") and kdv_amount is None:
                return (
                    "error",
                    None,
                    f"J{row} hücresindeki formülün hesaplanmış değeri okunamadı. "
                    "Dosyayı Excel'de açıp kaydederek yeniden yükleyin.",
                )
            if kdv_amount is None:
                continue

            current_base = _current_base_value(
                worksheet.cell(row, 9).value,
                values_worksheet.cell(row, 9).value,
                kdv_amount,
                row,
                workbook,
            )
            difference = None
            if current_base is not None:
                difference = (kdv_amount - (current_base * KDV_RATE)).quantize(
                    MONEY_STEP,
                    rounding=ROUND_HALF_UP,
                )

            if difference is None or difference != Decimal("0.00"):
                final_base = (kdv_amount / KDV_RATE).quantize(
                    MONEY_STEP,
                    rounding=ROUND_HALF_UP,
                )
                corrected_count += 1
            else:
                final_base = current_base.quantize(MONEY_STEP, rounding=ROUND_HALF_UP)

            worksheet.cell(row, 9).value = float(final_base)
            worksheet.cell(row, 10).value = float(kdv_amount)

            worksheet.cell(row, 13).value = f"=I{row}*0.20"
            worksheet.cell(row, 14).value = f"=ROUND(J{row}-M{row},2)"
            worksheet.cell(row, 13).number_format = "#,##0.00"
            worksheet.cell(row, 14).number_format = "#,##0.00"
            total_base += final_base
            total_kdv += kdv_amount
            processed_count += 1

        if processed_count == 0:
            return "error", None, "'Sonuç' sayfasında sayısal KDV tutarı bulunan veri satırı yok."

        worksheet["M1"] = None
        worksheet["N1"] = None

        if total_row is not None:
            total_base_cell = worksheet.cell(total_row, 9)
            if isinstance(total_base_cell.value, str) and total_base_cell.value.startswith("="):
                total_base_cell.value = float(total_base.quantize(MONEY_STEP, rounding=ROUND_HALF_UP))

            total_kdv_cell = worksheet.cell(total_row, 10)
            if isinstance(total_kdv_cell.value, str) and total_kdv_cell.value.startswith("="):
                cached_total = _as_decimal(values_worksheet.cell(total_row, 10).value)
                final_total = cached_total if cached_total is not None else total_kdv
                total_kdv_cell.value = float(final_total.quantize(MONEY_STEP, rounding=ROUND_HALF_UP))

        worksheet.column_dimensions["M"].width = 11.7109375
        worksheet.column_dimensions["N"].width = 13

        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"

        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(
            output_dir,
            f"ithaldeindirilecekfinal_{uuid.uuid4().hex[:8]}.xlsx",
        )
        workbook.save(output_path)
        return (
            "success",
            output_path,
            f"{processed_count} satır işlendi; I ve J sütunları değer olarak sabitlendi; "
            f"{corrected_count} satırın KDV matrahı düzeltildi.",
        )
    except Exception as exc:
        return "error", None, f"Excel dosyası işlenirken hata oluştu: {exc}"
    finally:
        if workbook is not None:
            workbook.close()
        if values_workbook is not None:
            values_workbook.close()


def _find_last_data_row(worksheet) -> int:
    last_data_row = 0
    for row in range(2, worksheet.max_row + 1):
        values = [worksheet.cell(row, column).value for column in range(1, 15)]
        if any(isinstance(value, str) and value.strip().upper() == "TOPLAM" for value in values):
            break
        if any(value not in (None, "") for value in values):
            last_data_row = row
    return last_data_row


def _find_total_row(worksheet):
    for row in range(2, worksheet.max_row + 1):
        values = [worksheet.cell(row, column).value for column in range(1, 15)]
        if any(isinstance(value, str) and value.strip().upper() == "TOPLAM" for value in values):
            return row
    return None


def _current_base_value(value, cached_value, kdv_amount: Decimal, row: int, workbook):
    numeric_value = _as_decimal(value)
    if numeric_value is not None:
        return numeric_value

    if isinstance(value, str) and value.startswith("="):
        normalized = value.replace(" ", "").upper()
        direct_formulas = {f"=J{row}/0.2", f"=J{row}/0.20", f"=J{row}/20%"}
        if normalized in direct_formulas:
            return kdv_amount / KDV_RATE

        normalized_control_sheet = CONTROL_SHEET.upper().replace(" ", "")
        control_formula = f"=J{row}/'{normalized_control_sheet}'!$B$2"
        if normalized == control_formula and CONTROL_SHEET in workbook.sheetnames:
            workbook_rate = _as_decimal(workbook[CONTROL_SHEET]["B2"].value)
            if workbook_rate not in (None, Decimal("0")):
                return kdv_amount / workbook_rate

    return _as_decimal(cached_value)


def _as_decimal(value):
    if isinstance(value, bool) or value is None:
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
