import os
import re
import unicodedata
import uuid
import xml.etree.ElementTree as ET
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


INVOICE_HEADER_ALIASES = {
    "faturaninno",
    "faturaninnosu",
    "faturanumarasi",
    "faturano",
    "faturanosu",
}
INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")
WORD_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
WORD_TABLE = f"{{{WORD_NS['w']}}}tbl"


def _normalize_text(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value).casefold())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "", text)


def _cell_text(cell) -> str:
    paragraphs = []
    for paragraph in cell.findall(".//w:p", WORD_NS):
        text = "".join(node.text or "" for node in paragraph.findall(".//w:t", WORD_NS))
        if text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


def _invoice_column(cells) -> int | None:
    for index, cell in enumerate(cells):
        if _normalize_text(_cell_text(cell)) in INVOICE_HEADER_ALIASES:
            return index
    return None


def _invoice_values(cell_text: str) -> list[str]:
    values = []
    for line in re.split(r"[\r\n]+", cell_text):
        value = re.sub(r"\s+", "", line)
        if value and _normalize_text(value) not in INVOICE_HEADER_ALIASES:
            values.append(value)
    return values


def extract_invoice_numbers(docx_path: str) -> list[str]:
    """Belgedeki FATURANIN NOSU sütunlarının tüm veri satırlarını döndürür."""
    with zipfile.ZipFile(docx_path) as archive:
        document_xml = archive.read("word/document.xml")

    document = ET.fromstring(document_xml)
    invoice_numbers = []

    for table in document.iter(WORD_TABLE):
        rows = table.findall("./w:tr", WORD_NS)
        for header_row_index, row in enumerate(rows):
            cells = row.findall("./w:tc", WORD_NS)
            invoice_col = _invoice_column(cells)
            if invoice_col is None:
                continue

            for data_row in rows[header_row_index + 1:]:
                data_cells = data_row.findall("./w:tc", WORD_NS)
                if invoice_col >= len(data_cells):
                    continue
                invoice_numbers.extend(_invoice_values(_cell_text(data_cells[invoice_col])))
            break

    return invoice_numbers


def _sheet_name(filename: str, used_names: set[str]) -> str:
    stem = os.path.splitext(os.path.basename(filename))[0]
    base = INVALID_SHEET_CHARS.sub("_", stem).strip().strip("'") or "Belge"
    base = base[:31]
    candidate = base
    number = 2

    while candidate.casefold() in used_names:
        suffix = f" ({number})"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        number += 1

    used_names.add(candidate.casefold())
    return candidate


def _format_sheet(worksheet, invoice_numbers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    divider = Side(style="thin", color="D9E2F3")

    worksheet["A1"] = "Fatura No"
    worksheet["A1"].fill = header_fill
    worksheet["A1"].font = header_font
    worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    worksheet["A1"].border = Border(bottom=divider)
    worksheet.row_dimensions[1].height = 22

    for row_index, invoice_number in enumerate(invoice_numbers, start=2):
        cell = worksheet.cell(row_index, 1, invoice_number)
        cell.font = body_font
        cell.number_format = "@"
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(bottom=divider)

    longest = max([len("Fatura No"), *(len(value) for value in invoice_numbers)])
    worksheet.column_dimensions["A"].width = min(max(longest + 3, 18), 45)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:A{max(len(invoice_numbers) + 1, 1)}"
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.tabColor = "1F4E78"


def word_invoices_to_excel(documents: list[tuple[str, str]], output_dir: str) -> tuple[str, str, str]:
    """Her Word belgesi için ayrı sayfada fatura numaraları bulunan Excel üretir."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_names = set()
    warnings = []
    total_invoice_count = 0

    for path, original_filename in documents:
        invoice_numbers = []
        try:
            invoice_numbers = extract_invoice_numbers(path)
        except Exception as exc:
            warnings.append(f"{original_filename}: Word belgesi okunamadı ({exc}).")
        else:
            if not invoice_numbers:
                warnings.append(
                    f"{original_filename}: 'FATURANIN NOSU' sütununda kayıt bulunamadı."
                )

        worksheet = workbook.create_sheet(_sheet_name(original_filename, used_names))
        _format_sheet(worksheet, invoice_numbers)
        total_invoice_count += len(invoice_numbers)

    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(
        output_dir,
        f"word_fatura_nolari_{uuid.uuid4().hex[:8]}.xlsx",
    )
    workbook.save(output_path)
    workbook.close()

    message = (
        f"Tamamlandı. {len(documents)} Word belgesinden "
        f"{total_invoice_count} fatura numarası aktarıldı."
    )
    if warnings:
        return "partial", output_path, f"{message} {' | '.join(warnings)}"
    return "success", output_path, message
