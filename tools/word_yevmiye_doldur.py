import os
import re
import shutil
import unicodedata
import uuid
import zipfile
from datetime import date, datetime
from xml.dom import Node, minidom

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from tools.common import secure_tr_filename


EXCEL_HEADERS = {
    "record_date": {"kayittarihi"},
    "document_date": {"belgetarihi"},
    "journal_no": {"yevmiyeno", "yevmiye numarasi"},
    "reference": {"referans", "faturano", "faturanosu"},
}
WORD_HEADERS = {
    "invoice": {"faturaninnosu", "faturano", "faturanosu", "faturanumarasi"},
    "record_date": {"yevmiyekayittarih", "yevmiyekayittarihi"},
    "voucher_no": {"mahsupfisno", "mahsupfisnosu", "mahsupfisnumarasi"},
}


def _normalize_text(value) -> str:
    text = str(value or "").casefold().replace("ı", "i")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "", text)


def _identifier(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\s+", "", str(value).strip())


def _journal_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _date_value(value, epoch) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value, epoch)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            return None
    if isinstance(value, str):
        text = value.strip()
        for pattern in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, pattern).date()
            except ValueError:
                continue
    return None


def _find_excel_columns(worksheet) -> tuple[int, dict[str, int]]:
    last_header_row = min(worksheet.max_row, 20)
    for row_number in range(1, last_header_row + 1):
        normalized = {
            _normalize_text(worksheet.cell(row_number, col).value): col
            for col in range(1, worksheet.max_column + 1)
            if worksheet.cell(row_number, col).value is not None
        }
        resolved = {}
        for key, aliases in EXCEL_HEADERS.items():
            for alias in aliases:
                column = normalized.get(_normalize_text(alias))
                if column:
                    resolved[key] = column
                    break
        if len(resolved) == len(EXCEL_HEADERS):
            return row_number, resolved

    expected = "Kayıt tarihi, Belge tarihi, Yevmiye No ve Referans"
    raise ValueError(f"Excel dosyasında gerekli başlıklar bulunamadı: {expected}.")


def read_excel_records(excel_path: str) -> dict[str, dict]:
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    try:
        worksheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        header_row, columns = _find_excel_columns(worksheet)
        grouped = {}

        for row_number in range(header_row + 1, worksheet.max_row + 1):
            reference = _identifier(worksheet.cell(row_number, columns["reference"]).value)
            if not reference:
                continue

            key = reference.casefold()
            item = grouped.setdefault(
                key,
                {
                    "reference": reference,
                    "record_dates": set(),
                    "document_dates": set(),
                    "journal_numbers": set(),
                },
            )
            item["record_dates"].add(
                _date_value(worksheet.cell(row_number, columns["record_date"]).value, workbook.epoch)
            )
            item["document_dates"].add(
                _date_value(worksheet.cell(row_number, columns["document_date"]).value, workbook.epoch)
            )
            item["journal_numbers"].add(
                _journal_text(worksheet.cell(row_number, columns["journal_no"]).value)
            )

        records = {}
        for key, item in grouped.items():
            record_dates = item["record_dates"]
            document_dates = item["document_dates"]
            journal_numbers = item["journal_numbers"]

            resolved_date = None
            if (
                len(record_dates) == 1
                and len(document_dates) == 1
                and None not in record_dates
                and None not in document_dates
                and record_dates == document_dates
            ):
                resolved_date = next(iter(record_dates)).strftime("%d.%m.%Y")

            resolved_journal = None
            if len(journal_numbers) == 1 and "" not in journal_numbers:
                resolved_journal = next(iter(journal_numbers))

            records[key] = {
                "reference": item["reference"],
                "record_date": resolved_date,
                "journal_no": resolved_journal,
            }
        return records
    finally:
        workbook.close()


def _direct_elements(parent, tag_name: str):
    return [
        child
        for child in parent.childNodes
        if child.nodeType == Node.ELEMENT_NODE and child.tagName == tag_name
    ]


def _cell_text(cell) -> str:
    return "".join(
        node.firstChild.data
        for node in cell.getElementsByTagName("w:t")
        if node.firstChild is not None
    )


def _word_columns(cells) -> dict[str, int] | None:
    normalized = {_normalize_text(_cell_text(cell)): index for index, cell in enumerate(cells)}
    resolved = {}
    for key, aliases in WORD_HEADERS.items():
        for alias in aliases:
            column = normalized.get(_normalize_text(alias))
            if column is not None:
                resolved[key] = column
                break
    return resolved if len(resolved) == len(WORD_HEADERS) else None


def _clean_run(document, run):
    for child in list(run.childNodes):
        if child.nodeType != Node.ELEMENT_NODE or child.tagName != "w:rPr":
            run.removeChild(child)
    return run


def _replace_cell_text(document, cell, value: str, source_cell) -> None:
    paragraphs = _direct_elements(cell, "w:p")
    if paragraphs:
        paragraph = paragraphs[0]
    else:
        paragraph = document.createElement("w:p")
        cell.appendChild(paragraph)

    existing_runs = paragraph.getElementsByTagName("w:r")
    source_runs = source_cell.getElementsByTagName("w:r")
    if existing_runs:
        run = _clean_run(document, existing_runs[0].cloneNode(deep=True))
    elif source_runs:
        run = _clean_run(document, source_runs[0].cloneNode(deep=True))
    else:
        run = document.createElement("w:r")

    for child in list(paragraph.childNodes):
        if child.nodeType != Node.ELEMENT_NODE or child.tagName != "w:pPr":
            paragraph.removeChild(child)

    text_node = document.createElement("w:t")
    text_node.appendChild(document.createTextNode(value))
    run.appendChild(text_node)
    paragraph.appendChild(run)

    for extra_paragraph in paragraphs[1:]:
        cell.removeChild(extra_paragraph)


def _write_docx(input_path: str, output_path: str, document_xml: bytes) -> None:
    with zipfile.ZipFile(input_path, "r") as source, zipfile.ZipFile(output_path, "w") as target:
        for item in source.infolist():
            data = document_xml if item.filename == "word/document.xml" else source.read(item.filename)
            target.writestr(item, data)


def fill_word_journal_table(
    word_path: str,
    excel_path: str,
    output_path: str,
) -> dict:
    records = read_excel_records(excel_path)
    with zipfile.ZipFile(word_path, "r") as archive:
        document_xml = archive.read("word/document.xml")

    document = minidom.parseString(document_xml)
    target_table_found = False
    invoice_count = 0
    filled_date_count = 0
    filled_journal_count = 0
    attention_invoices = set()

    for table in document.getElementsByTagName("w:tbl"):
        rows = _direct_elements(table, "w:tr")
        for header_index, header_row in enumerate(rows):
            header_cells = _direct_elements(header_row, "w:tc")
            columns = _word_columns(header_cells)
            if not columns:
                continue

            target_table_found = True
            for row in rows[header_index + 1:]:
                cells = _direct_elements(row, "w:tc")
                if max(columns.values()) >= len(cells):
                    continue

                invoice_number = _identifier(_cell_text(cells[columns["invoice"]]))
                if not invoice_number:
                    continue

                invoice_count += 1
                record = records.get(invoice_number.casefold())
                if record and record["record_date"]:
                    _replace_cell_text(
                        document,
                        cells[columns["record_date"]],
                        record["record_date"],
                        cells[columns["invoice"]],
                    )
                    filled_date_count += 1
                else:
                    attention_invoices.add(invoice_number)

                if record and record["journal_no"]:
                    _replace_cell_text(
                        document,
                        cells[columns["voucher_no"]],
                        record["journal_no"],
                        cells[columns["invoice"]],
                    )
                    filled_journal_count += 1
                else:
                    attention_invoices.add(invoice_number)
            break

    if not target_table_found:
        raise ValueError(
            "Word belgesinde FATURANIN NOSU, YEVMİYE KAYIT TARİHİ ve MAHSUP FİŞ NO başlıklı tablo bulunamadı."
        )
    if invoice_count == 0:
        raise ValueError("Word tablosunda işlenecek fatura numarası bulunamadı.")

    _write_docx(word_path, output_path, document.toxml(encoding="utf-8"))
    return {
        "invoice_count": invoice_count,
        "filled_date_count": filled_date_count,
        "filled_journal_count": filled_journal_count,
        "attention_invoices": sorted(attention_invoices),
    }


def process_word_journal_pairs(
    pairs: list[tuple[str, str, str]],
    output_dir: str,
) -> tuple[str, str, str]:
    os.makedirs(output_dir, exist_ok=True)
    batch_id = uuid.uuid4().hex[:8]
    completed = []
    warnings = []
    total_invoices = 0
    attention_document_count = 0

    for word_path, excel_path, original_word_name in pairs:
        stem = os.path.splitext(os.path.basename(original_word_name))[0]
        safe_name = secure_tr_filename(original_word_name)
        physical_path = os.path.join(output_dir, f"{batch_id}_{uuid.uuid4().hex[:6]}_{safe_name}")

        try:
            result = fill_word_journal_table(word_path, excel_path, physical_path)
            total_invoices += result["invoice_count"]
            needs_attention = bool(result["attention_invoices"])
            if needs_attention:
                attention_document_count += 1
                archive_name = f"{stem} (ilgilen).docx"
                warnings.append(
                    f"{original_word_name}: boş bırakılan faturalar: "
                    f"{', '.join(result['attention_invoices'])}"
                )
            else:
                archive_name = f"{stem}.docx"
        except Exception as exc:
            attention_document_count += 1
            archive_name = f"{stem} (ilgilen).docx"
            shutil.copyfile(word_path, physical_path)
            warnings.append(f"{original_word_name}: işlenemedi ({exc})")

        completed.append((physical_path, archive_name))

    zip_path = os.path.join(output_dir, f"word_yevmiye_sonuclari_{batch_id}.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for path, archive_name in completed:
            archive.write(path, arcname=archive_name)

    message = (
        f"Tamamlandı. {len(pairs)} Word belgesinde {total_invoices} fatura satırı işlendi."
    )
    if warnings:
        message = f"{message} {' | '.join(warnings)}"
        return "partial", zip_path, message
    return "success", zip_path, message
