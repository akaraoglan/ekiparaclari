import os
import re
import unicodedata
import uuid
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel


KDV_ORANI = Decimal("0.20")
VERGI_NUMARASI = 1111111111
PARA_BIRIMI_FORMATI = '#,##0.00'

MUAVIN_COLUMNS = {
    "date": ["Belge tarihi"],
    "document_title": ["Belge başlığı metni"],
    "amount": ["Şirket kodu para birimi değeri"],
    "text": ["Metin"],
}

IMPORT_COLUMNS = {
    "invoice": ["Fatura numarası"],
    "supplier": ["Tedarikçi adı"],
    "description": ["Mal grubu tanımı"],
    "quantity": ["Teslimat miktarı"],
    "unit": ["SAS ölçü birimi"],
    "declaration": ["Gümrük beyanname numarası"],
    "declaration_date": ["Gümrük beyanname tarihi"],
}

RESULT_HEADERS = [
    "SIRA NO",
    "TARİH",
    "SERİ NO",
    "FATURA NO",
    "CARİ HESAP ADI",
    "VERGİ NO",
    "FATURA İZAHAT",
    "MİKTAR",
    "KDV MATRAHI",
    "KDV TUTARI",
    "GGB",
    "KDV DÖNEMİ",
]

EXCEPTION_HEADERS = [
    "MUAVİN SATIRI",
    "BELGE TARİHİ",
    "BEYANNAME NO",
    "MUAVİN KDV TUTARI",
    "İTHALAT RAPORU SATIRI",
    "FATURA NO",
    "TEDARİKÇİ",
    "MAL GRUBU",
    "MİKTAR",
    "BİRİM",
    "NOT",
]


def process_ithalde_indirilecek_kdv(muavin_path: str, ithalat_raporu_path: str, output_dir: str) -> tuple:
    """
    Muavindeki ithalat KDV ödemelerini ana kaynak kabul ederek İthalde
    İndirilecek KDV listesini hazırlar.

    Döner: (status, output_path, message)
    """
    try:
        payments, tahakkuk_total = _read_muavin(muavin_path)
        import_by_declaration = _read_import_report(ithalat_raporu_path)

        result_rows = []
        exception_rows = []
        exception_declarations = set()

        for payment in payments:
            declaration = payment["declaration"]
            if not declaration:
                exception_declarations.add(payment["group_key"])
                exception_rows.append(_exception_row(
                    payment,
                    note="Muavinde beyanname numarası bulunamadı; sonuç satırı oluşturulmadı.",
                ))
                continue

            source_rows = import_by_declaration.get(declaration)
            if not source_rows:
                exception_declarations.add(declaration)
                exception_rows.append(_exception_row(
                    payment,
                    note="Muavindeki beyanname numarası ithalat raporunda bulunamadı; sonuç satırı oluşturulmadı.",
                ))
                continue

            problem = _validate_source_rows(source_rows)
            if problem:
                exception_declarations.add(declaration)
                for index, source in enumerate(source_rows):
                    exception_rows.append(_exception_row(
                        payment,
                        source=source,
                        include_amount=index == 0,
                        note=problem,
                    ))
                continue

            invoice_groups = _group_import_rows(source_rows)
            total_quantity = sum(group["quantity"] for group in invoice_groups)
            if total_quantity <= 0:
                exception_declarations.add(declaration)
                for index, source in enumerate(source_rows):
                    exception_rows.append(_exception_row(
                        payment,
                        source=source,
                        include_amount=index == 0,
                        note="Toplam miktar sıfır veya negatif olduğu için KDV dağıtımı yapılmadı.",
                    ))
                continue

            allocated = Decimal("0")
            for index, group in enumerate(invoice_groups):
                if index == len(invoice_groups) - 1:
                    vat_amount = payment["amount"] - allocated
                else:
                    vat_amount = _round_two(payment["amount"] * group["quantity"] / total_quantity)
                    allocated += vat_amount

                result_rows.append({
                    "date": payment["date"],
                    "invoice": group["invoice"],
                    "supplier": group["supplier"],
                    "description": group["description"],
                    "quantity": group["quantity"],
                    "unit": group["unit"],
                    "vat": vat_amount,
                    "declaration": declaration,
                    "period": _period_from_date(payment["date"]),
                })

        result_total = sum((row["vat"] for row in result_rows), Decimal("0"))
        exception_total = sum(
            (row["amount"] for row in exception_rows if row["amount"] is not None),
            Decimal("0"),
        )
        payment_total = sum((payment["amount"] for payment in payments), Decimal("0"))

        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(
            output_dir,
            f"ithalde_indirilecek_kdv_{uuid.uuid4().hex[:8]}.xlsx",
        )
        _write_workbook(
            result_rows=result_rows,
            exception_rows=exception_rows,
            tahakkuk_total=tahakkuk_total,
            payment_total=payment_total,
            result_total=result_total,
            exception_total=exception_total,
            output_path=output_path,
        )

        tahakkuk_difference = _round_two(tahakkuk_total - payment_total)
        message = (
            f"Tamamlandı. {len(result_rows)} sonuç satırı oluşturuldu. "
            f"Sonuç KDV toplamı: {_format_tr_money(result_total)} TL."
        )
        warnings = []
        if exception_declarations:
            warnings.append(
                f"{len(exception_declarations)} kayıt/beyanname İnceleme Gerekenler sayfasına ayrıldı "
                f"({_format_tr_money(exception_total)} TL)."
            )
        if abs(tahakkuk_difference) > Decimal("0.01"):
            warnings.append(
                "Muavin ödeme toplamı ile tahakkuk arasında "
                f"{_format_tr_money(tahakkuk_difference)} TL fark var."
            )

        if warnings:
            return "partial", output_path, f"{message} {' '.join(warnings)}"
        return "success", output_path, message

    except Exception as exc:
        return "error", None, f"Hata oluştu: {exc}"


def _read_muavin(path: str) -> tuple:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        columns = _resolve_columns(worksheet, MUAVIN_COLUMNS, "Muavin")
        raw_payments = []
        tahakkuk_total = Decimal("0")

        for row_num, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            document_title = _clean_text(values[columns["document_title"] - 1])
            row_text = _clean_text(values[columns["text"] - 1])
            combined = f"{document_title} {row_text}".upper()

            if "TAHAKKUK" in combined and "KDV" in combined:
                amount = _decimal_value(values[columns["amount"] - 1], row_num, "KDV tahakkuku")
                if amount:
                    tahakkuk_total += abs(amount)
                continue

            if "İTHALAT KDV ÖD" not in combined:
                continue

            amount = _decimal_value(values[columns["amount"] - 1], row_num, "KDV tutarı")
            raw_payments.append({
                "row_num": row_num,
                "date": _excel_date(values[columns["date"] - 1]),
                "declaration": _extract_declaration(row_text),
                "amount": amount,
                "raw_text": row_text,
            })

        if not raw_payments:
            raise ValueError("Muavinde 'İTHALAT KDV ÖD.' satırı bulunamadı.")
        if tahakkuk_total == 0:
            raise ValueError("Muavinde KDV tahakkuk kaydı bulunamadı veya tahakkuk tutarı sıfır.")

        grouped = {}
        order = []
        for item in raw_payments:
            key = item["declaration"] or f"__BEYANNAME_YOK_{item['row_num']}"
            if key not in grouped:
                grouped[key] = {
                    "group_key": key,
                    "declaration": item["declaration"],
                    "date": item["date"],
                    "amount": Decimal("0"),
                    "row_nums": [],
                    "raw_texts": [],
                }
                order.append(key)
            grouped[key]["amount"] += item["amount"]
            grouped[key]["row_nums"].append(item["row_num"])
            grouped[key]["raw_texts"].append(item["raw_text"])

        return [grouped[key] for key in order], _round_two(tahakkuk_total)
    finally:
        workbook.close()


def _read_import_report(path: str) -> dict:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        columns = _resolve_columns(worksheet, IMPORT_COLUMNS, "İthalat Raporu")
        rows_by_declaration = defaultdict(list)

        for row_num, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            declaration = _normalize_declaration(values[columns["declaration"] - 1])
            if not declaration:
                continue
            rows_by_declaration[declaration].append({
                "row_num": row_num,
                "invoice_raw": _clean_text(values[columns["invoice"] - 1]),
                "invoice": _normalize_invoice(values[columns["invoice"] - 1]),
                "supplier": _clean_text(values[columns["supplier"] - 1]),
                "description": _clean_text(values[columns["description"] - 1]),
                "quantity": _decimal_value(values[columns["quantity"] - 1], row_num, "Teslimat miktarı"),
                "unit": _clean_text(values[columns["unit"] - 1]).upper(),
                "declaration": declaration,
                "declaration_date": values[columns["declaration_date"] - 1],
            })

        if not rows_by_declaration:
            raise ValueError("İthalat Raporu'nda beyanname numarası bulunan satır yok.")
        return rows_by_declaration
    finally:
        workbook.close()


def _validate_source_rows(source_rows: list):
    if any(not row["invoice"] for row in source_rows):
        return "İthalat raporunda fatura numarası boş olduğu için yazılmadı."
    if any(not row["unit"] for row in source_rows):
        return "İthalat raporunda miktar birimi boş olduğu için yazılmadı."

    units = {row["unit"] for row in source_rows}
    if len(units) > 1:
        return "Miktar birimleri farklı olduğu için yazılmadı."

    by_invoice = defaultdict(list)
    for row in source_rows:
        by_invoice[row["invoice"]].append(row)
    for rows in by_invoice.values():
        suppliers = {_normalize_comparison_text(row["supplier"]) for row in rows}
        if len(suppliers) > 1:
            return "Aynı fatura için tedarikçi adları farklı olduğu için yazılmadı."
    return None


def _group_import_rows(source_rows: list) -> list:
    groups = {}
    order = []
    for row in source_rows:
        key = row["invoice"]
        if key not in groups:
            groups[key] = {
                "invoice": key,
                "supplier": row["supplier"],
                "description": row["description"],
                "descriptions": [],
                "quantity": Decimal("0"),
                "unit": row["unit"],
            }
            order.append(key)
        groups[key]["quantity"] += row["quantity"]
        if row["description"] and row["description"] not in groups[key]["descriptions"]:
            groups[key]["descriptions"].append(row["description"])
        groups[key]["description"] = " / ".join(groups[key]["descriptions"])
    return [groups[key] for key in order]


def _exception_row(payment: dict, source=None, include_amount=True, note="") -> dict:
    source = source or {}
    return {
        "muavin_rows": ", ".join(str(row) for row in payment["row_nums"]),
        "date": payment["date"],
        "declaration": payment["declaration"],
        "amount": payment["amount"] if include_amount else None,
        "import_row": source.get("row_num"),
        "invoice": source.get("invoice_raw", ""),
        "supplier": source.get("supplier", ""),
        "description": source.get("description", ""),
        "quantity": source.get("quantity"),
        "unit": source.get("unit", ""),
        "note": note,
    }


def _write_workbook(
    result_rows: list,
    exception_rows: list,
    tahakkuk_total: Decimal,
    payment_total: Decimal,
    result_total: Decimal,
    exception_total: Decimal,
    output_path: str,
) -> None:
    workbook = Workbook()
    result_sheet = workbook.active
    result_sheet.title = "Sonuç"
    exception_sheet = workbook.create_sheet("İnceleme Gerekenler")
    control_sheet = workbook.create_sheet("Kontrol Özeti")

    _write_result_sheet(result_sheet, result_rows)
    _write_exception_sheet(exception_sheet, exception_rows)
    _write_control_sheet(
        control_sheet,
        result_sheet.max_row,
        exception_sheet.max_row,
        bool(exception_rows),
        tahakkuk_total,
        payment_total,
        result_total,
        exception_total,
    )

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.save(output_path)


def _write_result_sheet(ws, rows: list) -> None:
    header_fill = PatternFill("solid", fgColor="FFF200")
    row_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(name="Arial", size=10, bold=True)
    data_font = Font(name="Arial", size=9)
    border = Border(bottom=Side(style="thin", color="A6B1BC"))

    ws.append(RESULT_HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color="7F6000"))
    ws.row_dimensions[1].height = 32

    for index, item in enumerate(rows, start=1):
        row_num = index + 1
        ws.append([
            index,
            item["date"],
            "",
            item["invoice"],
            item["supplier"],
            VERGI_NUMARASI,
            item["description"],
            f"{_format_quantity(item['quantity'])}{item['unit']}",
            None,
            float(item["vat"]),
            item["declaration"],
            item["period"],
        ])
        ws.cell(row_num, 9, f"=J{row_num}/'Kontrol Özeti'!$B$2")
        for cell in ws[row_num]:
            cell.fill = row_fill
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        for col in (1, 2, 3, 6, 8, 9, 10, 11, 12):
            ws.cell(row_num, col).alignment = Alignment(horizontal="center", vertical="center")

    total_row = len(rows) + 2
    ws.cell(total_row, 9, "TOPLAM")
    ws.cell(total_row, 10, f"=SUM(J2:J{total_row - 1})" if rows else "=0")
    for col in (9, 10):
        ws.cell(total_row, col).font = Font(name="Arial", size=10, bold=True)
        ws.cell(total_row, col).fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(total_row, col).border = Border(top=Side(style="double", color="7F6000"))

    widths = [10, 13, 11, 23, 39, 16, 24, 16, 18, 18, 25, 15]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{max(1, total_row - 1)}"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["D"].number_format = "@"
    ws.column_dimensions["K"].number_format = "@"
    for row in range(2, total_row + 1):
        ws.cell(row, 2).number_format = "dd.mm.yyyy"
        ws.cell(row, 4).number_format = "@"
        ws.cell(row, 8).number_format = "@"
        ws.cell(row, 9).number_format = PARA_BIRIMI_FORMATI
        ws.cell(row, 10).number_format = PARA_BIRIMI_FORMATI
        ws.cell(row, 11).number_format = "@"
        ws.cell(row, 12).number_format = "0"


def _write_exception_sheet(ws, rows: list) -> None:
    header_fill = PatternFill("solid", fgColor="F4B183")
    exception_fill = PatternFill("solid", fgColor="FFF2CC")
    header_font = Font(name="Arial", size=10, bold=True)
    data_font = Font(name="Arial", size=9)

    ws.append(EXCEPTION_HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color="C65911"))
    ws.row_dimensions[1].height = 34

    for item in rows:
        ws.append([
            item["muavin_rows"],
            item["date"],
            item["declaration"],
            float(item["amount"]) if item["amount"] is not None else None,
            item["import_row"],
            item["invoice"],
            item["supplier"],
            item["description"],
            float(item["quantity"]) if item["quantity"] is not None else None,
            item["unit"],
            item["note"],
        ])
        for cell in ws[ws.max_row]:
            cell.fill = exception_fill
            cell.font = data_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="thin", color="D6B656"))

    widths = [15, 14, 26, 20, 21, 25, 40, 25, 14, 12, 65]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{max(1, ws.max_row)}"
    ws.sheet_view.showGridLines = False
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 2).number_format = "dd.mm.yyyy"
        ws.cell(row, 3).number_format = "@"
        ws.cell(row, 4).number_format = PARA_BIRIMI_FORMATI
        ws.cell(row, 6).number_format = "@"
        ws.cell(row, 9).number_format = '#,##0.###'


def _write_control_sheet(
    ws,
    result_last_row: int,
    exception_last_row: int,
    has_exceptions: bool,
    tahakkuk_total: Decimal,
    payment_total: Decimal,
    result_total: Decimal,
    exception_total: Decimal,
) -> None:
    status = "HAZIR"
    if abs(tahakkuk_total - payment_total) > Decimal("0.01"):
        status = "TAHAKKUK UYUMSUZ"
    elif has_exceptions:
        status = "ELLE TAMAMLAMA GEREKİYOR"

    ws.append(["KONTROL", "DEĞER", "AÇIKLAMA"])
    rows = [
        ["KDV Oranı", float(KDV_ORANI), "KDV Matrahı = KDV Tutarı ÷ KDV Oranı"],
        ["Muavin Tahakkuk Toplamı", float(tahakkuk_total), "Muavindeki KDV tahakkuk satırı/satırları"],
        ["Muavin İthalat KDV Ödeme Toplamı", float(payment_total), "Muavindeki İTHALAT KDV ÖD. satırları"],
        ["Sonuç KDV Toplamı", f"='Sonuç'!J{result_last_row}", "Ana sonuç sayfasına yazılan tutar"],
        ["İnceleme Gereken KDV", f"=SUM('İnceleme Gerekenler'!D2:D{max(2, exception_last_row)})", "Elle tamamlanacak kayıtların tutarı"],
        ["Sonuç + İnceleme", "=B5+B6", "Muavin ödeme toplamına eşit olmalıdır"],
        ["Ödeme Kontrol Farkı", "=ROUND(B4-B7,2)", "Sıfır olmalıdır"],
        ["Tahakkuk Kontrol Farkı", "=ROUND(B3-B4,2)", "Sıfır olmalıdır"],
        ["DURUM", status, "İstisna varsa İnceleme Gerekenler sayfasını kontrol edin"],
    ]
    for row in rows:
        ws.append(row)

    for row in range(1, ws.max_row + 1):
        for cell in ws[row]:
            cell.font = Font(name="Arial", size=10, bold=row in (1, 10))
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="thin", color="D9E2F3"))
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    ws["B10"].fill = PatternFill(
        "solid",
        fgColor="C6E0B4" if status == "HAZIR" else "F4B183",
    )
    ws["B10"].font = Font(name="Arial", size=10, bold=True)
    ws["B2"].number_format = "0%"
    for row in range(3, 10):
        ws.cell(row, 2).number_format = PARA_BIRIMI_FORMATI
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 58
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _resolve_columns(ws, required: dict, label: str) -> dict:
    normalized_headers = {}
    header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    for col, value in enumerate(header_values, start=1):
        if value is not None:
            normalized_headers[_normalize_header(value)] = col

    resolved = {}
    missing = []
    for key, aliases in required.items():
        column = next(
            (normalized_headers.get(_normalize_header(alias)) for alias in aliases
             if normalized_headers.get(_normalize_header(alias))),
            None,
        )
        if column:
            resolved[key] = column
        else:
            missing.append(aliases[0])
    if missing:
        raise ValueError(f"{label} dosyasında eksik kolon: {', '.join(missing)}")
    return resolved


def _normalize_header(value) -> str:
    text = unicodedata.normalize("NFKD", str(value).strip().casefold())
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"[^a-z0-9]+", "", text)


def _normalize_declaration(value) -> str:
    return re.sub(r"\s+", "", _clean_text(value).upper())


def _normalize_invoice(value) -> str:
    return re.sub(r"[^A-Z0-9]+", "", _clean_text(value).upper())


def _normalize_comparison_text(value) -> str:
    return re.sub(r"\s+", " ", _clean_text(value).upper()).strip()


def _extract_declaration(value) -> str:
    match = re.search(r"\(([^()]*)\)", _clean_text(value))
    if not match:
        return ""
    return _normalize_declaration(match.group(1))


def _decimal_value(value, row_num: int, field_name: str) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip().replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"{row_num}. satırda {field_name} sayısal değil: {value}") from exc


def _excel_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        converted = from_excel(value)
        return converted.date() if isinstance(converted, datetime) else converted
    if value:
        text = str(value).strip()
        for pattern in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, pattern).date()
            except ValueError:
                pass
    raise ValueError(f"Geçersiz belge tarihi: {value}")


def _period_from_date(value: date) -> int:
    return value.year * 100 + value.month


def _clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _round_two(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _format_quantity(value: Decimal) -> str:
    normalized = value.normalize()
    if normalized == normalized.to_integral():
        return str(int(normalized))
    return format(normalized, "f")


def _format_tr_money(value: Decimal) -> str:
    value = _round_two(value)
    formatted = f"{value:,.2f}"
    return formatted.replace(",", "_").replace(".", ",").replace("_", ".")
