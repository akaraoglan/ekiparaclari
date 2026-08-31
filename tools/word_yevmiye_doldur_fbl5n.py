import os
import shutil
import uuid
import zipfile

from openpyxl import load_workbook

from tools.common import secure_tr_filename
from tools.word_yevmiye_doldur import (
    _cell_text,
    _date_value,
    _direct_elements,
    _identifier,
    _normalize_text,
    _replace_cell_text,
    _word_columns,
    _write_docx,
)


EXCEL_HEADERS = {
    "reference": {"referans"},
    "document_no": {"belgenumarasi", "belgeno", "belgenosu"},
    "record_date": {"kayittarihi"},
    "document_type": {"belgeturu"},
    "clearing_no": {
        "denklestirmebelgesi",
        "denklestirmenumarasi",
        "denklestirmeno",
    },
}


def _resolve_excel_columns(row_values) -> dict[str, int] | None:
    normalized = {
        _normalize_text(value): index
        for index, value in enumerate(row_values)
        if value is not None
    }
    resolved = {}
    for key, aliases in EXCEL_HEADERS.items():
        for alias in aliases:
            column = normalized.get(_normalize_text(alias))
            if column is not None:
                resolved[key] = column
                break
    return resolved if len(resolved) == len(EXCEL_HEADERS) else None


def _document_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_fbl5n_records(excel_path: str) -> dict[str, dict]:
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    try:
        worksheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        columns = None
        source_rows = []
        reversed_clearing_numbers = set()

        for row_number, row_values in enumerate(
            worksheet.iter_rows(values_only=True),
            start=1,
        ):
            if columns is None:
                if row_number <= 20:
                    columns = _resolve_excel_columns(row_values)
                    if columns is not None:
                        continue
                if row_number >= 20:
                    expected = (
                        "Referans, Belge numarası, Kayıt tarihi, Belge türü ve "
                        "Denkleştirme belgesi"
                    )
                    raise ValueError(
                        f"Excel dosyasında gerekli başlıklar bulunamadı: {expected}."
                    )
                continue

            reference = _identifier(row_values[columns["reference"]])
            document_type = _normalize_text(row_values[columns["document_type"]])
            clearing_no = _identifier(row_values[columns["clearing_no"]])
            source_rows.append(
                {
                    "reference": reference,
                    "document_no": _document_text(row_values[columns["document_no"]]),
                    "record_date": _date_value(
                        row_values[columns["record_date"]],
                        workbook.epoch,
                    ),
                    "document_type": document_type,
                    "clearing_no": clearing_no,
                }
            )
            if document_type == "tk" and clearing_no:
                reversed_clearing_numbers.add(clearing_no.casefold())

        if columns is None:
            expected = (
                "Referans, Belge numarası, Kayıt tarihi, Belge türü ve "
                "Denkleştirme belgesi"
            )
            raise ValueError(f"Excel dosyasında gerekli başlıklar bulunamadı: {expected}.")

        grouped = {}
        for row in source_rows:
            if not row["reference"]:
                continue

            clearing_key = row["clearing_no"].casefold()
            if row["document_type"] == "tk" or (
                clearing_key and clearing_key in reversed_clearing_numbers
            ):
                continue

            reference_key = row["reference"].casefold()
            item = grouped.setdefault(
                reference_key,
                {
                    "reference": row["reference"],
                    "record_dates": set(),
                    "document_numbers": set(),
                },
            )
            item["record_dates"].add(row["record_date"])
            item["document_numbers"].add(row["document_no"])

        records = {}
        for key, item in grouped.items():
            record_dates = item["record_dates"]
            document_numbers = item["document_numbers"]

            resolved_date = None
            if len(record_dates) == 1 and None not in record_dates:
                resolved_date = next(iter(record_dates)).strftime("%d.%m.%Y")

            resolved_document_no = None
            if len(document_numbers) == 1 and "" not in document_numbers:
                resolved_document_no = next(iter(document_numbers))

            records[key] = {
                "reference": item["reference"],
                "record_date": resolved_date,
                "document_no": resolved_document_no,
            }
        return records
    finally:
        workbook.close()


def fill_word_journal_table_fbl5n(
    word_path: str,
    excel_path: str,
    output_path: str,
) -> dict:
    records = read_fbl5n_records(excel_path)
    with zipfile.ZipFile(word_path, "r") as archive:
        document_xml = archive.read("word/document.xml")

    from xml.dom import minidom

    document = minidom.parseString(document_xml)
    target_table_found = False
    invoice_count = 0
    filled_date_count = 0
    filled_document_count = 0
    attention_invoices = set()

    for table in document.getElementsByTagName("w:tbl"):
        rows = _direct_elements(table, "w:tr")
        for header_index, header_row in enumerate(rows):
            header_cells = _direct_elements(header_row, "w:tc")
            columns = _word_columns(header_cells)
            if not columns:
                continue

            target_table_found = True
            for row in rows[header_index + 1:]:
                cells = _direct_elements(row, "w:tc")
                if max(columns.values()) >= len(cells):
                    continue

                invoice_number = _identifier(_cell_text(cells[columns["invoice"]]))
                if not invoice_number:
                    continue

                invoice_count += 1
                record = records.get(invoice_number.casefold())
                if record and record["record_date"]:
                    _replace_cell_text(
                        document,
                        cells[columns["record_date"]],
                        record["record_date"],
                        cells[columns["invoice"]],
                    )
                    filled_date_count += 1
                else:
                    attention_invoices.add(invoice_number)

                if record and record["document_no"]:
                    _replace_cell_text(
                        document,
                        cells[columns["voucher_no"]],
                        record["document_no"],
                        cells[columns["invoice"]],
                    )
                    filled_document_count += 1
                else:
                    attention_invoices.add(invoice_number)
            break

    if not target_table_found:
        raise ValueError(
            "Word belgesinde FATURANIN NOSU, YEVMİYE KAYIT TARİHİ ve "
            "MAHSUP FİŞ NO başlıklı tablo bulunamadı."
        )
    if invoice_count == 0:
        raise ValueError("Word tablosunda işlenecek fatura numarası bulunamadı.")

    _write_docx(word_path, output_path, document.toxml(encoding="utf-8"))
    return {
        "invoice_count": invoice_count,
        "filled_date_count": filled_date_count,
        "filled_document_count": filled_document_count,
        "attention_invoices": sorted(attention_invoices),
    }


def process_word_journal_pairs_fbl5n(
    pairs: list[tuple[str, str, str]],
    output_dir: str,
) -> tuple[str, str, str]:
    os.makedirs(output_dir, exist_ok=True)
    batch_id = uuid.uuid4().hex[:8]
    completed = []
    warnings = []
    total_invoices = 0

    for word_path, excel_path, original_word_name in pairs:
        stem = os.path.splitext(os.path.basename(original_word_name))[0]
        safe_name = secure_tr_filename(original_word_name)
        physical_path = os.path.join(
            output_dir,
            f"{batch_id}_{uuid.uuid4().hex[:6]}_{safe_name}",
        )

        try:
            result = fill_word_journal_table_fbl5n(word_path, excel_path, physical_path)
            total_invoices += result["invoice_count"]
            if result["attention_invoices"]:
                archive_name = f"{stem} (ilgilen).docx"
                warnings.append(
                    f"{original_word_name}: boş bırakılan faturalar: "
                    f"{', '.join(result['attention_invoices'])}"
                )
            else:
                archive_name = f"{stem}.docx"
        except Exception as exc:
            archive_name = f"{stem} (ilgilen).docx"
            shutil.copyfile(word_path, physical_path)
            warnings.append(f"{original_word_name}: işlenemedi ({exc})")

        completed.append((physical_path, archive_name))

    zip_path = os.path.join(output_dir, f"word_yevmiye_fbl5n_sonuclari_{batch_id}.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for path, archive_name in completed:
            archive.write(path, arcname=archive_name)

    message = (
        f"Tamamlandı. {len(pairs)} Word belgesinde {total_invoices} fatura satırı işlendi."
    )
    if warnings:
        message = f"{message} {' | '.join(warnings)}"
        return "partial", zip_path, message
    return "success", zip_path, message
