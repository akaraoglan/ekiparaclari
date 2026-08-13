import os
import re
import uuid
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side


KDV_RATE = Decimal("0.20")
MONEY_STEP = Decimal("0.01")
OFFICIAL_WIDTHS = {
    "A": 6.28515625,
    "B": 4.7109375,
    "C": 12.5703125,
    "D": 11.42578125,
    "E": 21.42578125,
    "F": 64.7109375,
    "G": 21.5703125,
    "H": 17.85546875,
    "I": 10.140625,
    "J": 6.85546875,
    "K": 13.85546875,
    "L": 13.140625,
    "M": 15.42578125,
    "N": 21.28515625,
    "O": 15.28515625,
    "P": 9.140625,
}
OFFICIAL_NUMBER_FORMATS = {
    "B": "0",
    "C": "mm-dd-yy",
    "D": "@",
    "E": "0",
    "F": "@",
    "G": "@",
    "H": "0",
    "I": "#,##0.00",
    "J": "@",
    "K": "#,##0.00",
    "L": "#,##0.00",
    "M": "mm-dd-yy",
    "N": "@",
    "O": "@",
}
OFFICIAL_ALIGNMENTS = {
    "B": "center",
    "C": "center",
    "D": "center",
    "E": "center",
    "F": "left",
    "G": "center",
    "H": "center",
    "I": "right",
    "J": "center",
    "K": "right",
    "L": "right",
    "M": "center",
    "N": "center",
    "O": "center",
}


def process_ihrac_kayitli_final(input_path: str, output_dir: str) -> tuple:
    """İhraç kayıtlı sonuç dosyasındaki matrah ve KDV farklarını kesinleştirir."""
    workbook = None
    values_workbook = None
    try:
        workbook = load_workbook(input_path, data_only=False)
        values_workbook = load_workbook(input_path, data_only=True)
    except Exception:
        if workbook is not None:
            workbook.close()
        if values_workbook is not None:
            values_workbook.close()
        return "error", None, "Excel dosyası açılamadı. Lütfen geçerli bir .xlsx dosyası yükleyin."

    try:
        sheet_name = _find_result_sheet(workbook.sheetnames)
        if sheet_name is None:
            return "error", None, "Excel dosyasında 'Sonuc' veya 'Sonuç' sayfası bulunamadı."

        worksheet = workbook[sheet_name]
        values_worksheet = values_workbook[sheet_name]
        header_row = _find_header_row(worksheet)
        if header_row is None:
            return "error", None, "K ve L sütunlarındaki matrah/KDV başlıkları bulunamadı."

        total_row = _find_total_row(worksheet, header_row)
        if total_row is None:
            return "error", None, "K sütununun altında 'TOPLAM' satırı bulunamadı."

        rows = []
        total_kdv_from_rows = Decimal("0")
        for row in range(header_row + 1, total_row):
            kdv = _cell_decimal(worksheet, values_worksheet, row, 12)
            base_value = _cell_decimal(worksheet, values_worksheet, row, 11)
            if kdv is None and base_value is None:
                continue
            if kdv is None:
                return "error", None, f"L{row} hücresinde sayısal KDV tutarı bulunamadı."

            kdv = kdv.quantize(MONEY_STEP, rounding=ROUND_HALF_UP)
            rows.append((row, base_value, kdv))
            total_kdv_from_rows += kdv

        if not rows:
            return "error", None, "İşlenecek sayısal KDV satırı bulunamadı."

        total_kdv = _cell_decimal(worksheet, values_worksheet, total_row, 12)
        if total_kdv is None:
            total_kdv = total_kdv_from_rows
        total_kdv = total_kdv.quantize(MONEY_STEP, rounding=ROUND_HALF_UP)
        total_kdv_from_rows = total_kdv_from_rows.quantize(MONEY_STEP, rounding=ROUND_HALF_UP)
        if total_kdv != total_kdv_from_rows:
            return (
                "error",
                None,
                f"L sütunundaki satır toplamı ({_tr_money(total_kdv_from_rows)}) ile "
                f"L{total_row} toplamı ({_tr_money(total_kdv)}) eşleşmiyor.",
            )

        target_base_total = (total_kdv / KDV_RATE).quantize(MONEY_STEP, rounding=ROUND_HALF_UP)
        changed_count = 0
        final_base_total = Decimal("0")

        for row, current_base, kdv in rows:
            final_base = (kdv / KDV_RATE).quantize(MONEY_STEP, rounding=ROUND_HALF_UP)
            current_rounded = None
            if current_base is not None:
                current_rounded = current_base.quantize(MONEY_STEP, rounding=ROUND_HALF_UP)
            if current_rounded != final_base:
                changed_count += 1

            worksheet.cell(row, 11).value = float(final_base)
            worksheet.cell(row, 11).number_format = "#,##0.00"
            calculated_kdv = final_base * KDV_RATE

            difference = kdv - calculated_kdv
            if difference != Decimal("0"):
                return "error", None, f"Q{row} farkı 0,00 değerine getirilemedi."
            final_base_total += final_base

        final_base_total = final_base_total.quantize(MONEY_STEP, rounding=ROUND_HALF_UP)
        if final_base_total != target_base_total:
            return (
                "error",
                None,
                f"Matrah toplamı {_tr_money(final_base_total)} yerine "
                f"{_tr_money(target_base_total)} olmalıdır.",
            )

        _apply_official_layout(
            worksheet,
            header_row,
            total_row,
            [row for row, _, _ in rows],
            final_base_total,
            total_kdv,
        )

        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"

        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(
            output_dir,
            f"ihrac_kayitli_final_{uuid.uuid4().hex[:8]}.xlsx",
        )
        workbook.save(output_path)
        return (
            "success",
            output_path,
            f"{len(rows)} satır işlendi; {changed_count} matrah düzeltildi. "
            f"Tüm KDV farkları 0,00 ve matrah toplamı {_tr_money(target_base_total)} olarak doğrulandı.",
        )
    except Exception as exc:
        return "error", None, f"Excel dosyası işlenirken hata oluştu: {exc}"
    finally:
        if workbook is not None:
            workbook.close()
        if values_workbook is not None:
            values_workbook.close()


def _find_result_sheet(sheet_names):
    for name in sheet_names:
        if _normalize_text(name) == "sonuc":
            return name
    return None


def _find_header_row(worksheet):
    for row in range(1, worksheet.max_row + 1):
        k_header = _normalize_text(worksheet.cell(row, 11).value)
        l_header = _normalize_text(worksheet.cell(row, 12).value)
        if "kdvharictutari" in k_header and "kdvsi" in l_header:
            return row
    return None


def _find_total_row(worksheet, header_row):
    for row in range(header_row + 1, worksheet.max_row + 1):
        if _normalize_text(worksheet.cell(row, 11).value) == "toplam":
            return row
    return None


def _cell_decimal(worksheet, values_worksheet, row, column):
    value = worksheet.cell(row, column).value
    numeric = _as_decimal(value)
    if numeric is not None:
        return numeric
    if isinstance(value, str) and value.startswith("="):
        return _as_decimal(values_worksheet.cell(row, column).value)
    return None


def _normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower().translate(str.maketrans({
        "ı": "i",
        "ğ": "g",
        "ü": "u",
        "ş": "s",
        "ö": "o",
        "ç": "c",
    }))
    return re.sub(r"[^a-z0-9]+", "", text)


def _as_decimal(value):
    if value is None or isinstance(value, bool):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _tr_money(value: Decimal) -> str:
    formatted = f"{value:,.2f}"
    return formatted.replace(",", "_").replace(".", ",").replace("_", ".")


def _apply_official_layout(
    worksheet,
    header_row: int,
    total_row: int,
    data_rows: list,
    total_base: Decimal,
    total_kdv: Decimal,
) -> None:
    """Resmî İnternet Vergi Dairesi şablonunun ölçü ve hücre biçimlerini uygular."""
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name="Arial", size=14, bold=True)
    header_font = Font(name="Arial", size=10, bold=True)
    data_font = Font(name="Arial", size=9)
    total_font = Font(name="Arial", size=9, bold=True)

    for column, width in OFFICIAL_WIDTHS.items():
        worksheet.column_dimensions[column].width = width
    if "Q" in worksheet.column_dimensions:
        del worksheet.column_dimensions["Q"]

    worksheet.sheet_format.defaultRowHeight = 15
    worksheet.sheet_format.baseColWidth = 8
    worksheet.row_dimensions[2].height = 18
    worksheet.row_dimensions[header_row].height = 140.25
    for row in data_rows + [total_row]:
        worksheet.row_dimensions[row].height = None

    title = worksheet["H2"]
    title.font = title_font
    title.alignment = Alignment(horizontal="center", vertical="center")
    title.number_format = "@"

    for column in range(2, 16):
        cell = worksheet.cell(header_row, column)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        cell.number_format = OFFICIAL_NUMBER_FORMATS[cell.column_letter]

    for row in data_rows:
        for column in range(2, 16):
            cell = worksheet.cell(row, column)
            cell.font = data_font
            cell.alignment = Alignment(
                horizontal=OFFICIAL_ALIGNMENTS[cell.column_letter],
                vertical="center",
            )
            cell.border = border
            cell.number_format = OFFICIAL_NUMBER_FORMATS[cell.column_letter]

    worksheet.cell(total_row, 11).value = "TOPLAM"
    worksheet.cell(total_row, 12).value = float(total_base)
    worksheet.cell(total_row, 13).value = float(total_kdv)
    for column in range(11, 14):
        cell = worksheet.cell(total_row, column)
        cell.font = total_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.number_format = "#,##0.00"

    for column in range(14, worksheet.max_column + 1):
        worksheet.cell(total_row, column).value = None
    if worksheet.max_column > 15:
        worksheet.delete_cols(16, worksheet.max_column - 15)

    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = "9"
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_setup.fitToWidth = None
    worksheet.page_margins.left = 0
    worksheet.page_margins.right = 0
    worksheet.page_margins.top = 0.984251968503937
    worksheet.page_margins.bottom = 0.984251968503937
    worksheet.page_margins.header = 0.5118110236220472
    worksheet.page_margins.footer = 0.5118110236220472
